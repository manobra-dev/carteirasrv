#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════
  CARTEIRAS TEMÁTICAS — SERVIDOR LOCAL DO DASHBOARD
═══════════════════════════════════════════════════════════

Uso:
    python3 serve.py                   → busca dados + abre o browser
    python3 serve.py --porta 8080      → porta customizada
    python3 serve.py --sem-browser     → não abre o browser automaticamente
    python3 serve.py --inicio 2024-01  → data inicial custom (YYYY-MM)
    python3 serve.py --sem-fundamentus → pula busca de DY do Fundamentus
    python3 serve.py --sem-yfinance    → usa brapi.dev para preços (sem retorno total)

O que faz:
  1. Busca cotações mensais AJUSTADAS por dividendos via yfinance (retorno total)
     ↳ Fallback: brapi.dev (somente retorno de preço, sem dividendos)
  2. Busca Dividend Yield via Fundamentus.com.br (mais completo)
     ↳ Fallback: brapi.dev → valores estáticos
  3. Salva em dashboard_data.json (lido pelo HTML sem CORS)
  4. Inicia servidor HTTP local na pasta atual
  5. Abre o browser em http://localhost:{porta}/dashboard_carteiras_v2.html

Por que yfinance para preços?
  O preço de fechamento bruto NÃO inclui dividendos. Quando a empresa paga
  dividendo, o preço cai (ex-dividendo) mas o investidor recebeu o dinheiro.
  O yfinance com auto_adjust=True retroage os preços históricos para refletir
  o retorno total (preço + dividendos reinvestidos), que é o número correto
  para comparar desempenho de carteiras focadas em dividendos.

Por que é necessário?
  Ao abrir o HTML diretamente (file://), o browser bloqueia
  requisições a APIs externas por segurança (CORS). Servindo
  via localhost:// esse bloqueio não se aplica. Além disso,
  o JSON local elimina dependência de rede ao visualizar.
"""

import argparse
import http.server
import json
import os
import sys
import time
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

import requests
import subprocess

# ─── Configuração ─────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
OUTPUT_JSON = SCRIPT_DIR / "dashboard_data.json"

# ── Estado global do refresh ──────────────────────────────
_refresh_state = {
    "running":   False,
    "last_ok":   None,   # ISO timestamp do último refresh bem-sucedido
    "error":     None,   # mensagem do último erro (ou None)
}
_refresh_lock = threading.Lock()

# Parâmetros de execução (preenchidos no __main__ e usados pelo auto-refresh)
_run_params: dict = {}

# Railway injeta a variável PORT automaticamente; localmente usa 8080
DEFAULT_PORT = int(os.environ.get("PORT", 8080))

# True quando rodando no Railway (ou qualquer ambiente de nuvem)
IS_CLOUD = bool(os.environ.get("RAILWAY_ENVIRONMENT") or
                os.environ.get("RENDER") or
                os.environ.get("DYNO"))  # Heroku

API_BASE = "https://brapi.dev/api"

# ── Token brapi.dev (opcional, mas recomendado) ────────────
# Cole seu token abaixo entre as aspas.
# Obtenha gratuitamente em: https://brapi.dev/dashboard
# Deixe None para usar sem token (sujeito a rate-limit).
BRAPI_TOKEN = "j8ymSoxSAvAp53ULGbFCgN"

TICKER_ALIAS = {
    "EMBJ3": "EMBR3",
    "AXIA3": "ELET3",
    "AXIA6": "ELET6",
}

# Tickers que trocaram de código na B3/BDR em uma data específica.
# A série histórica é "costurada": código antigo até o mês de corte,
# código novo a partir do mês de corte (inclusive).
TICKER_STITCH = {
    # JBS: ação B3 (JBBS3) até Jun/2025 → BDR (JBSS32) a partir de Jul/2025
    "JBSS32": {"old": "JBBS3", "new": "JBSS32", "cutoff": "2025-07"},
    # Embraer: EMBR3 até Set/2025 → EMBJ3 a partir de Out/2025 (renomeação na B3)
    "EMBJ3":  {"old": "EMBR3", "new": "EMBJ3",  "cutoff": "2025-10"},
}

PORTFOLIOS = {
    "acoes": {
        "name": "Carteira de Ações",
        "ativos": {
            "BPAC11": {"peso": 0.05, "setor": "Financeiro"},
            "BBAS3":  {"peso": 0.05, "setor": "Financeiro"},
            "ITUB4":  {"peso": 0.12, "setor": "Financeiro"},
            "JPMC34": {"peso": 0.05, "setor": "Financeiro"},
            "CPLE3":  {"peso": 0.08, "setor": "Energia"},
            "EMBJ3":  {"peso": 0.10, "setor": "Aviação"},
            "WEGE3":  {"peso": 0.05, "setor": "Indústria"},
            "GGBR4":  {"peso": 0.05, "setor": "Siderurgia"},
            "VALE3":  {"peso": 0.10, "setor": "Commodities"},
            "AURA33": {"peso": 0.10, "setor": "Mineração"},
            "AXIA3":  {"peso": 0.05, "setor": "Energia Elétrica"},
            "TTEN3":  {"peso": 0.06, "setor": "Agroindústria"},
            "CSMG3":  {"peso": 0.04, "setor": "Saneamento"},
            "IVVB11": {"peso": 0.10, "setor": "Índices"},
        },
    },
    "dividendos": {
        "name": "Carteira de Dividendos",
        "ativos": {
            "BBDC4":  {"peso": 0.10, "setor": "Financeiro"},
            "CXSE3":  {"peso": 0.08, "setor": "Seguros"},
            "ITUB4":  {"peso": 0.15, "setor": "Financeiro"},
            "B3SA3":  {"peso": 0.05, "setor": "Financeiro"},
            "SANB11": {"peso": 0.05, "setor": "Financeiro"},
            "NDIV11": {"peso": 0.10, "setor": "Índice"},
            "PETR4":  {"peso": 0.12, "setor": "Petroquímica"},
            "VALE3":  {"peso": 0.10, "setor": "Commodities"},
            "CSMG3":  {"peso": 0.15, "setor": "Saneamento"},
            "AXIA6":  {"peso": 0.10, "setor": "Energia Elétrica"},
        },
    },
}

DY_FALLBACK = {
    # ── Carteira de Ações (ativos atuais) ─────────────────
    "BPAC11": 2.1,   # BTG Pactual
    "BBAS3":  8.8,   # Banco do Brasil
    "ITUB4":  4.2,   # Itaú PN (inclui JCP)
    "JPMC34": 2.8,   # BDR JPMorgan
    "CPLE3":  6.1,   # Copel ON
    "EMBJ3":  0.9,   # Embraer
    "WEGE3":  1.5,   # WEG
    "GGBR4":  5.3,   # Gerdau PN
    "VALE3":  10.2,  # Vale (dividendo variável)
    "AURA33": 2.0,   # Aura Minerals
    "AXIA3":  7.1,   # Eletrobras ON
    "TTEN3":  2.3,   # Tten
    "CSMG3":  5.4,   # Copasa
    "IVVB11": 1.2,   # ETF S&P 500
    # ── Carteira de Dividendos (ativos atuais) ─────────────
    "BBDC4":  6.2,   # Bradesco PN (inclui JCP)
    "CXSE3":  5.1,   # Caixa Seguridade
    "B3SA3":  5.0,   # B3
    "SANB11": 6.5,   # Santander Brasil units (inclui JCP)
    "NDIV11": 7.2,   # ETF IDIV
    "PETR4":  12.8,  # Petrobras PN (variável)
    "AXIA6":  7.4,   # Eletrobras PNB
    "CPFE3":  5.8,   # CPFL Energia
    # ── Ativos históricos (já saíram da carteira) ──────────
    "JBSS32": 3.8,   # JBS S.A. BDR (jul/2025+)
    "JBBS3":  3.8,   # JBS S.A. B3 (até jun/2025)
    "BPAN4":  3.2,   # Banco Pan PN
    "SUZB3":  4.5,   # Suzano
    "CPLE6":  6.4,   # Copel PNB
    "IGTI11": 2.8,   # Iguatemi units
    "ARZZ3":  1.9,   # Arezzo (agora AZZA3)
    "AZZA3":  2.1,   # Azzas 2154 (ex-Arezzo)
    "SMAL11": 3.5,   # ETF Small Caps
    "CPLE5":  6.2,   # Copel PNA
    "BBSE3":  5.9,   # BB Seguridade
}

# ─── Fetcher ──────────────────────────────────────────────

class DataFetcher:

    HEADERS = {"User-Agent": "CarteirasAnalytics/2.0"}
    BATCH   = 5
    RETRIES = 3

    def __init__(self, token: str | None = None):
        self.session = requests.Session()
        self.session.headers.update(self.HEADERS)
        # Prioridade: argumento CLI → constante BRAPI_TOKEN no topo do arquivo
        self.token = token or BRAPI_TOKEN

    def _params(self, extra: dict) -> dict:
        p = dict(extra)
        if self.token:
            p["token"] = self.token
        return p

    def _get(self, ticker_str: str, interval="1mo", range_="2y", fundamental=False) -> dict | None:
        params = self._params({"range": range_, "interval": interval})
        if fundamental:
            params["fundamental"] = "true"

        for attempt in range(self.RETRIES):
            try:
                r = self.session.get(f"{API_BASE}/quote/{ticker_str}", params=params, timeout=20)
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                wait = 2 ** attempt
                if attempt < self.RETRIES - 1:
                    print(f"    ⚠  Tentativa {attempt+1}/{self.RETRIES}: {e}. Aguardando {wait}s…")
                    time.sleep(wait)
                else:
                    print(f"    ✗  Falha definitiva: {ticker_str} → {e}")
        return None

    def _extract_monthly(self, hist: list, start_ts: int) -> dict[str, float]:
        """Extrai {YYYY-MM: close} de uma lista historicalDataPrice, sem duplicatas."""
        by_month: dict[str, tuple] = {}  # month_key → (close, orig_date)
        for d in hist:
            if not d.get("close") or d["date"] < start_ts:
                continue
            label = datetime.fromtimestamp(d["date"]).strftime("%Y-%m")
            # Mantém o registro com a data mais recente dentro do mês
            if label not in by_month or d["date"] > by_month[label][1]:
                by_month[label] = (round(d["close"], 4), d["date"])
        return {k: v[0] for k, v in by_month.items()}

    # ── yfinance (retorno total ajustado por dividendos) ─────

    def _ensure_yfinance(self):
        """Importa yfinance, instalando automaticamente se necessário."""
        try:
            import yfinance as yf
            return yf
        except ImportError:
            print("    ℹ  Instalando 'yfinance' (necessário para retorno total)…")
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "yfinance", "-q"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
                import yfinance as yf
                print("    ✓  'yfinance' instalado com sucesso.")
                return yf
            except Exception as e:
                print(f"    ⚠  Não foi possível instalar 'yfinance': {e}")
                return None

    def _yf_symbol(self, display_ticker: str) -> str:
        """Converte ticker de exibição para símbolo Yahoo Finance (.SA para B3/BDR).
        NÃO aplica TICKER_ALIAS — Yahoo Finance já usa os tickers atuais (EMBJ3, AXIA3, AXIA6).
        TICKER_ALIAS é exclusivo do brapi.dev (que ainda indexa pelos tickers antigos)."""
        return display_ticker + ".SA"

    def fetch_prices_yfinance(
        self, display_tickers: list[str], start_dt: str
    ) -> dict[str, dict]:
        """
        Busca preços AJUSTADOS por dividendos via yfinance (retorno total).
        auto_adjust=True retroage os preços históricos descontando cada dividendo
        pago, de modo que a variação de preço já inclui o retorno dos dividendos.
        Retorna {display_ticker: {"YYYY-MM": adjusted_close, ...}}.
        """
        yf = self._ensure_yfinance()
        if yf is None:
            return {}

        prices: dict[str, dict] = {}

        # ── Tickers normais (batch) ───────────────────────
        stitch_set = {t for t in display_tickers if t in TICKER_STITCH}
        regular    = [t for t in display_tickers if t not in stitch_set]

        # yf_symbol → lista de display tickers que mapeiam nele
        yf_to_disp: dict[str, list[str]] = {}
        for t in regular:
            sym = self._yf_symbol(t)
            yf_to_disp.setdefault(sym, []).append(t)

        yf_symbols = list(yf_to_disp.keys())
        if yf_symbols:
            print(f"    📥  yfinance: {len(yf_symbols)} tickers (retorno total, ajustado por dividendos)…")
            try:
                import pandas as pd
                raw = yf.download(
                    yf_symbols, start=start_dt, interval="1mo",
                    auto_adjust=True, progress=False, group_by="ticker",
                )
                if raw.empty:
                    raise ValueError("DataFrame vazio")

                for sym, disp_list in yf_to_disp.items():
                    try:
                        # yfinance retorna MultiIndex quando >1 ticker
                        if len(yf_symbols) > 1:
                            col = (sym, "Close") if (sym, "Close") in raw.columns else None
                            if col is None:
                                # tenta sem o ticker prefix (yfinance v0.2 style)
                                series = raw["Close"].get(sym)
                            else:
                                series = raw[col]
                        else:
                            series = raw["Close"]
                        if series is None or (hasattr(series, "empty") and series.empty):
                            raise ValueError("série vazia")
                        series = series.dropna()
                        monthly = {
                            idx.strftime("%Y-%m"): round(float(v), 4)
                            for idx, v in series.items()
                            if not pd.isna(v)
                        }
                        if monthly:
                            for t in disp_list:
                                prices[t] = monthly
                            keys = sorted(monthly)
                            print(f"    ✓  {disp_list[0]:8s} ({sym:16s}): "
                                  f"{len(monthly)} meses [{keys[0]}→{keys[-1]}] [total return]")
                        else:
                            print(f"    ⚠  {sym}: sem dados yfinance")
                    except Exception as e:
                        print(f"    ⚠  {sym}: {e}")

            except Exception as e:
                print(f"    ⚠  yfinance batch falhou: {e}")

        # ── Tickers com troca de código (stitch via yfinance) ─
        for disp_t in stitch_set:
            cfg    = TICKER_STITCH[disp_t]
            old_sym = cfg["old"] + ".SA"
            new_sym = cfg["new"] + ".SA"
            cutoff  = cfg["cutoff"]
            print(f"\n  ↔  Costurando {disp_t} via yfinance: {old_sym}+{new_sym}")
            merged: dict[str, float] = {}
            try:
                import pandas as pd
                old_raw = yf.download(old_sym, start=start_dt, interval="1mo",
                                      auto_adjust=True, progress=False)
                if not old_raw.empty:
                    for idx, v in old_raw["Close"].dropna().items():
                        label = idx.strftime("%Y-%m")
                        if label < cutoff and not pd.isna(v):
                            merged[label] = round(float(v), 4)
            except Exception as e:
                print(f"    ⚠  {old_sym}: {e}")

            try:
                import pandas as pd
                new_raw = yf.download(new_sym, start=start_dt, interval="1mo",
                                      auto_adjust=True, progress=False)
                if not new_raw.empty:
                    for idx, v in new_raw["Close"].dropna().items():
                        label = idx.strftime("%Y-%m")
                        if label >= cutoff and not pd.isna(v):
                            merged[label] = round(float(v), 4)
            except Exception as e:
                print(f"    ⚠  {new_sym}: {e}")

            if merged:
                prices[disp_t] = merged
                keys  = sorted(merged)
                n_old = sum(1 for k in merged if k < cutoff)
                n_new = len(merged) - n_old
                print(f"    ✓  {disp_t} costurado: {len(merged)} meses "
                      f"[{keys[0]}→{keys[-1]}] "
                      f"({n_old}×{cfg['old']} + {n_new}×{cfg['new']}) [total return]")
            else:
                print(f"    ⚠  {disp_t}: sem dados após costura yfinance")

        return prices

    # ── Cotações DIÁRIAS via yfinance ─────────────────────

    def fetch_prices_daily_yfinance(
        self, display_tickers: list[str], start_dt: str
    ) -> dict[str, dict]:
        """
        Busca preços DIÁRIOS ajustados via yfinance.
        Retorna {display_ticker: {"YYYY-MM-DD": adjusted_close, ...}}.
        Usado apenas para os gráficos de rentabilidade acumulada (linha mais suave).
        """
        yf = self._ensure_yfinance()
        if yf is None:
            return {}

        prices: dict[str, dict] = {}

        stitch_set = {t for t in display_tickers if t in TICKER_STITCH}
        regular    = [t for t in display_tickers if t not in stitch_set]

        yf_to_disp: dict[str, list[str]] = {}
        for t in regular:
            sym = self._yf_symbol(t)
            yf_to_disp.setdefault(sym, []).append(t)

        yf_symbols = list(yf_to_disp.keys())
        if yf_symbols:
            print(f"    📥  yfinance diário: {len(yf_symbols)} tickers…")
            try:
                import pandas as pd
                raw = yf.download(
                    yf_symbols, start=start_dt, interval="1d",
                    auto_adjust=True, progress=False, group_by="ticker",
                )
                if raw.empty:
                    raise ValueError("DataFrame vazio")

                for sym, disp_list in yf_to_disp.items():
                    try:
                        if len(yf_symbols) > 1:
                            col = (sym, "Close") if (sym, "Close") in raw.columns else None
                            series = raw[col] if col else raw["Close"].get(sym)
                        else:
                            series = raw["Close"]
                        if series is None or (hasattr(series, "empty") and series.empty):
                            raise ValueError("série vazia")
                        series = series.dropna()
                        daily = {
                            idx.strftime("%Y-%m-%d"): round(float(v), 4)
                            for idx, v in series.items()
                            if not pd.isna(v)
                        }
                        if daily:
                            for t in disp_list:
                                prices[t] = daily
                            keys = sorted(daily)
                            print(f"    ✓  {disp_list[0]:8s} ({sym:16s}): "
                                  f"{len(daily)} dias [{keys[0]}→{keys[-1]}]")
                        else:
                            print(f"    ⚠  {sym}: sem dados diários")
                    except Exception as e:
                        print(f"    ⚠  {sym}: {e}")
            except Exception as e:
                print(f"    ⚠  yfinance diário batch falhou: {e}")

        # ── Tickers com troca de código (stitch diário) ───
        for disp_t in stitch_set:
            cfg     = TICKER_STITCH[disp_t]
            old_sym = cfg["old"] + ".SA"
            new_sym = cfg["new"] + ".SA"
            cutoff  = cfg["cutoff"]   # "YYYY-MM"
            print(f"\n  ↔  Costurando diário {disp_t}: {old_sym}+{new_sym}")
            merged: dict[str, float] = {}
            try:
                import pandas as pd
                old_raw = yf.download(old_sym, start=start_dt, interval="1d",
                                      auto_adjust=True, progress=False)
                if not old_raw.empty:
                    for idx, v in old_raw["Close"].dropna().items():
                        label = idx.strftime("%Y-%m-%d")
                        if label[:7] < cutoff and not pd.isna(v):
                            merged[label] = round(float(v), 4)
            except Exception as e:
                print(f"    ⚠  {old_sym}: {e}")
            try:
                import pandas as pd
                new_raw = yf.download(new_sym, start=start_dt, interval="1d",
                                      auto_adjust=True, progress=False)
                if not new_raw.empty:
                    for idx, v in new_raw["Close"].dropna().items():
                        label = idx.strftime("%Y-%m-%d")
                        if label[:7] >= cutoff and not pd.isna(v):
                            merged[label] = round(float(v), 4)
            except Exception as e:
                print(f"    ⚠  {new_sym}: {e}")

            if merged:
                prices[disp_t] = merged
                keys = sorted(merged)
                print(f"    ✓  {disp_t} diário costurado: {len(merged)} dias [{keys[0]}→{keys[-1]}]")
            else:
                print(f"    ⚠  {disp_t}: sem dados diários após costura")

        return prices

    # ── Preços: yfinance (total return) → brapi (price return) ──

    def fetch_prices(self, display_tickers: list[str], start_ts: int,
                     use_yfinance: bool = True) -> dict[str, dict]:
        """
        Retorna {display_ticker: {"YYYY-MM": close_price, ...}}
        Camada 1: yfinance auto_adjust=True  → retorno total (preço + dividendos)
        Camada 2: brapi.dev close            → retorno de preço apenas (fallback)
        Tickers em TICKER_STITCH são costurados dos dois códigos históricos.
        """
        prices: dict[str, dict] = {}
        start_dt = datetime.fromtimestamp(start_ts).strftime("%Y-%m-%d")

        # ── Camada 1: yfinance ────────────────────────────
        if use_yfinance:
            prices = self.fetch_prices_yfinance(display_tickers, start_dt)
            missing = [t for t in display_tickers if t not in prices]
            if missing:
                print(f"\n  ↩  Fallback brapi.dev para {len(missing)} ticker(s) "
                      f"sem dados yfinance: {missing}")
        else:
            missing = list(display_tickers)

        if not missing:
            return prices

        # ── Camada 2: brapi.dev (preço bruto, sem dividendos) ─
        stitch_set   = {t for t in missing if t in TICKER_STITCH}
        regular_list = [t for t in missing if t not in stitch_set]

        # ── Processamento normal (batch) ──────────────────
        for i in range(0, len(regular_list), self.BATCH):
            batch = regular_list[i:i + self.BATCH]
            # Resolve aliases e deduplica
            api_map = {t: TICKER_ALIAS.get(t, t) for t in batch}
            api_str = ",".join(dict.fromkeys(api_map.values()))  # preserva ordem, deduplica

            data = self._get(api_str)
            if not data or not data.get("results"):
                print(f"    ✗  Sem dados para batch: {batch}")
                time.sleep(0.5)
                continue

            for display_t, api_t in api_map.items():
                item = next((x for x in data["results"] if x["symbol"] == api_t), None)
                if not item:
                    print(f"    ⚠  {display_t} ({api_t}) não retornou dados")
                    continue
                hist = item.get("historicalDataPrice") or []
                monthly = self._extract_monthly(hist, start_ts)
                if monthly:
                    prices[display_t] = monthly
                    n = len(monthly)
                    keys = sorted(monthly)
                    print(f"    ✓  {display_t:8s} ({api_t:8s}): {n} meses  [{keys[0]} → {keys[-1]}]")
                else:
                    print(f"    ⚠  {display_t}: sem dados no período")

            time.sleep(0.2)  # cortesia rate-limit

        # ── Tickers com troca de código (stitch) ──────────
        for disp_t in stitch_set:
            cfg = TICKER_STITCH[disp_t]
            old_t, new_t, cutoff = cfg["old"], cfg["new"], cfg["cutoff"]
            print(f"\n  ↔  Costurando {disp_t}: {old_t} (até {cutoff}) + {new_t} ({cutoff}+)")

            merged: dict[str, float] = {}

            # Período antigo: código B3 (pré-cutoff)
            old_resp = self._get(old_t, interval="1mo", range_="5y")
            if old_resp and old_resp.get("results"):
                hist = old_resp["results"][0].get("historicalDataPrice") or []
                for label, price in self._extract_monthly(hist, start_ts).items():
                    if label < cutoff:
                        merged[label] = price

            time.sleep(0.2)

            # Período novo: BDR (cutoff em diante)
            new_resp = self._get(new_t, interval="1mo", range_="5y")
            if new_resp and new_resp.get("results"):
                hist = new_resp["results"][0].get("historicalDataPrice") or []
                for label, price in self._extract_monthly(hist, start_ts).items():
                    if label >= cutoff:
                        merged[label] = price

            if merged:
                prices[disp_t] = merged
                keys = sorted(merged)
                n_old = sum(1 for k in merged if k < cutoff)
                n_new = sum(1 for k in merged if k >= cutoff)
                print(f"    ✓  {disp_t} (costurado): {len(merged)} meses "
                      f"[{keys[0]} → {keys[-1]}]  "
                      f"({n_old}×{old_t} + {n_new}×{new_t})")
            else:
                print(f"    ⚠  {disp_t}: sem dados após costura")

            time.sleep(0.2)

        return prices

    # ── Dividend Yield: Fundamentus ──────────────────────────

    def _ensure_fundamentus(self):
        """Importa o pacote 'fundamentus', instalando se necessário."""
        try:
            import fundamentus
            return fundamentus
        except ImportError:
            print("    ℹ  Pacote 'fundamentus' não encontrado. Instalando…")
            import subprocess
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "fundamentus", "-q"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                import fundamentus
                print("    ✓  'fundamentus' instalado com sucesso.")
                return fundamentus
            except Exception as e:
                print(f"    ⚠  Não foi possível instalar 'fundamentus': {e}")
                return None

    def fetch_dy_fundamentus(self, tickers: list[str]) -> dict[str, float | None]:
        """
        Busca DY de todos os ativos de uma vez via Fundamentus.com.br.
        Retorna dict {ticker_display: dy_pct | None}.
        """
        fund = self._ensure_fundamentus()
        if fund is None:
            return {t: None for t in tickers}

        try:
            df = fund.get_resultado()
            # Normaliza índice para maiúsculas
            df.index = df.index.str.upper()

            # Encontra coluna de DY (pode ser 'dy' ou 'Div.Yield')
            dy_col = next(
                (c for c in df.columns
                 if c.lower() in ("dy", "div.yield", "dividend_yield", "dividendyield")),
                None
            )
            if dy_col is None:
                print(f"    ⚠  Coluna DY não encontrada no Fundamentus. Colunas: {df.columns.tolist()}")
                return {t: None for t in tickers}

            result = {}
            for t in tickers:
                api_t = TICKER_ALIAS.get(t, t).upper()
                if api_t in df.index:
                    raw = df.loc[api_t, dy_col]
                    try:
                        raw_f = float(raw)
                        # Fundamentus retorna decimal (0.085) → converte para %
                        val = round(raw_f * 100, 2) if raw_f < 1.0 else round(raw_f, 2)
                        result[t] = val if val > 0 else None
                    except (TypeError, ValueError):
                        result[t] = None
                else:
                    result[t] = None
            return result

        except Exception as e:
            print(f"    ⚠  Erro ao buscar dados do Fundamentus: {e}")
            return {t: None for t in tickers}

    def fetch_dy(self, tickers: list[str], use_fundamentus: bool = True) -> dict[str, float | None]:
        """
        Busca Dividend Yield com 3 camadas:
          1. Fundamentus.com.br  (bulk, mais completo e atualizado)
          2. brapi.dev           (fundamental=true, por ticker)
          3. DY_FALLBACK         (valores estáticos de referência)
        """
        fund_dy: dict[str, float | None] = {}

        # ── Camada 1: Fundamentus ──────────────────────────
        if use_fundamentus:
            print("    🔍  Buscando DY no Fundamentus.com.br…")
            fund_dy = self.fetch_dy_fundamentus(tickers)
            found = sum(1 for v in fund_dy.values() if v is not None)
            print(f"    ✓  Fundamentus: {found}/{len(tickers)} DYs obtidos")

        dy: dict[str, float | None] = {}

        for t in tickers:
            # Se Fundamentus retornou valor, usa direto
            if fund_dy.get(t) is not None:
                dy[t] = fund_dy[t]
                print(f"    DY {t:8s}: {dy[t]:.1f}%  [fundamentus]")
                continue

            # ── Camada 2: brapi.dev ────────────────────────
            api_t = TICKER_ALIAS.get(t, t)
            data = self._get(api_t, interval="1mo", range_="1y", fundamental=True)
            brapi_val = None
            if data and data.get("results"):
                raw = data["results"][0].get("dividendYield")
                # Ignora 0: brapi retorna 0 quando não tem dado, não DY real de 0%
                if raw is not None and raw > 0:
                    brapi_val = round(raw if raw > 1 else raw * 100, 2)

            if brapi_val is not None:
                dy[t] = brapi_val
                print(f"    DY {t:8s}: {dy[t]:.1f}%  [brapi]")
                time.sleep(0.1)
                continue

            # ── Camada 3: fallback estático ────────────────
            dy[t] = DY_FALLBACK.get(t)
            disp = f"{dy[t]:.1f}%" if dy[t] else "N/D"
            src  = "fallback" if dy[t] else "N/D"
            print(f"    DY {t:8s}: {disp:8s}  [{src}]")
            time.sleep(0.05)

        return dy


# ─── CDI Mensal (BCB) ─────────────────────────────────────

def fetch_cdi(start_date: datetime) -> dict[str, float]:
    """
    Busca CDI mensal acumulado via API do Banco Central do Brasil (série 12).
    Retorna {YYYY-MM: cdi_pct} onde cdi_pct é o retorno do CDI no mês em %.
    Ex.: {"2024-06": 0.79, "2024-07": 0.80, ...}
    """
    try:
        start_str = start_date.strftime("%d/%m/%Y")
        end_str   = datetime.now().strftime("%d/%m/%Y")
        url = (
            "https://api.bcb.gov.br/dados/serie/bcdata.sgs.12/dados"
            f"?formato=json&dataInicial={start_str}&dataFinal={end_str}"
        )
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()
        # Formato: [{"data": "01/06/2024", "valor": "0.79"}, ...]
        result = {}
        for item in data:
            parts = item["data"].split("/")          # DD/MM/YYYY
            month_key = f"{parts[2]}-{parts[1]}"    # YYYY-MM
            result[month_key] = round(float(item["valor"]), 6)
        print(f"    ✓  CDI: {len(result)} meses carregados do BCB")
        return result
    except Exception as e:
        print(f"    ⚠  Erro ao buscar CDI (BCB): {e}")
        return {}


# ─── Pesos históricos do Excel ────────────────────────────

_MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12,
}

def _parse_mes_col_header(val) -> str | None:
    """Converte cabeçalho de coluna (ex: 'Jun/2024') para 'YYYY-MM'. Retorna None se não for mês."""
    if not isinstance(val, str):
        return None
    val = val.strip()
    parts = val.split("/")
    if len(parts) == 2:
        mes_str, ano_str = parts[0].strip(), parts[1].strip()
        mes = _MESES_PT.get(mes_str.lower()[:3])
        if mes and ano_str.isdigit() and len(ano_str) == 4:
            return f"{ano_str}-{mes:02d}"
    return None


def ler_pesos_historicos(xlsx_path: Path) -> dict:
    """
    Lê o histórico de pesos mensais de cada carteira a partir do Excel de rebalanceamentos.
    Retorna:
        {
          "acoes":      {"2024-06": {"BPAC11": 7.0, "BBAS3": 10.0, ...}, ...},
          "dividendos": {"2024-06": {...}, ...},
        }
    Tickers com peso None/ausente naquele mês são ignorados (ativo não estava na carteira).
    """
    if not xlsx_path.exists():
        print(f"  ⚠  Excel de rebalanceamentos não encontrado: {xlsx_path}")
        return {}

    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        import openpyxl

    sheet_map = {"Ações": "acoes", "Dividendos": "dividendos"}
    result = {}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sheet_name, cart_key in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Encontra linha de cabeçalho (a que tem >= 2 colunas com formato mês/ano)
        header_idx = None
        for ri, row in enumerate(rows):
            meses = [_parse_mes_col_header(v) for v in row]
            if sum(1 for m in meses if m) >= 2:
                header_idx = ri
                break

        if header_idx is None:
            print(f"  ⚠  Cabeçalho de meses não encontrado na aba '{sheet_name}'")
            continue

        header = rows[header_idx]
        # Mapeia índice de coluna → "YYYY-MM"
        col_to_mes = {ci: _parse_mes_col_header(v) for ci, v in enumerate(header)}

        cart_hist: dict[str, dict] = {}  # {"2024-06": {"TICKER": peso}}

        for row in rows[header_idx + 1:]:
            if not row or row[0] is None:
                continue
            ticker = str(row[0]).strip()
            # Ignora linha de TOTAL ou instruções
            if ticker.upper() in ("TOTAL", "TICKER", "💡") or ticker.startswith("💡"):
                continue
            if not ticker or len(ticker) > 10:
                continue

            for ci, val in enumerate(row):
                mes = col_to_mes.get(ci)
                if not mes:
                    continue
                if val is None or val == "":
                    continue
                try:
                    peso = float(val)
                except (TypeError, ValueError):
                    continue
                if peso <= 0:
                    continue
                if mes not in cart_hist:
                    cart_hist[mes] = {}
                cart_hist[mes][ticker] = peso

        result[cart_key] = cart_hist
        n_meses = len(cart_hist)
        print(f"  ✓  Pesos históricos '{sheet_name}': {n_meses} meses carregados")

    return result


# ─── Build JSON ───────────────────────────────────────────

def build_dashboard_json(
    start_date: datetime,
    token: str | None = None,
    use_fundamentus: bool = True,
    use_yfinance: bool = True,
) -> dict:
    fetcher = DataFetcher(token=token)

    # ── Busca um mês antes do início para calcular retorno do 1º mês ──
    # Ex.: início = jun/2024 → busca preços a partir de mai/2024 como base.
    # O JSON guarda mai/2024, permitindo que o dashboard compute jun/2024
    # como (fech_jun - fech_mai) / fech_mai em vez de exibir "-".
    y, m = start_date.year, start_date.month
    m -= 1
    if m == 0:
        m, y = 12, y - 1
    fetch_start = datetime(y, m, 1)
    start_ts = int(fetch_start.timestamp())

    # ── Pesos históricos do Excel (lido PRIMEIRO para incluir novos ativos) ──
    xlsx_path = SCRIPT_DIR / "rebalanceamentos_template.xlsx"
    print(f"\n{'─'*54}")
    print("  Lendo pesos históricos do Excel…")
    print(f"{'─'*54}")
    weights_history = ler_pesos_historicos(xlsx_path)

    # Tickers base (hardcoded) + benchmarks
    all_display_tickers = list(dict.fromkeys(
        t
        for p in PORTFOLIOS.values()
        for t in p["ativos"]
    )) + ["BOVA11", "DIVO11"]

    # Adiciona tickers do Excel que ainda não estão na lista (novos ativos)
    extra_from_excel: set[str] = set()
    for port_hist in weights_history.values():
        for month_weights in port_hist.values():
            extra_from_excel.update(month_weights.keys())
    extra_from_excel -= set(all_display_tickers)
    if extra_from_excel:
        novos = sorted(extra_from_excel)
        print(f"  📋  Novos ativos no Excel (não hardcoded): {', '.join(novos)}")
        all_display_tickers.extend(novos)

    price_source = "yfinance (retorno total)" if use_yfinance else "brapi.dev (retorno de preço)"
    print(f"\n{'─'*54}")
    print(f"  Buscando cotações mensais ({len(all_display_tickers)} ativos)…")
    print(f"  Início: {start_date.strftime('%d/%m/%Y')}  (base: {fetch_start.strftime('%b/%Y')})  |  Fonte: {price_source}")
    print(f"{'─'*54}")

    prices = fetcher.fetch_prices(all_display_tickers, start_ts, use_yfinance=use_yfinance)

    # ── Cotações diárias (para gráficos de rentabilidade acumulada) ──
    prices_daily: dict = {}
    if use_yfinance:
        print(f"\n{'─'*54}")
        print(f"  Buscando cotações DIÁRIAS ({len(all_display_tickers)} ativos)…")
        print(f"{'─'*54}")
        prices_daily = fetcher.fetch_prices_daily_yfinance(
            all_display_tickers, fetch_start.strftime("%Y-%m-%d")
        )

    print(f"\n{'─'*54}")
    print("  Buscando Dividend Yield…")
    dy_source = "Fundamentus → brapi.dev → fallback" if use_fundamentus else "brapi.dev → fallback"
    print(f"  Fontes: {dy_source}")
    print(f"{'─'*54}")

    # Tickers sem benchmarks (ETFs de índice, DY não aplicável da mesma forma)
    dy_tickers = [t for t in all_display_tickers if t not in ("BOVA11", "DIVO11")]
    dy = fetcher.fetch_dy(dy_tickers, use_fundamentus=use_fundamentus)

    # ── CDI mensal (BCB) ──────────────────────────────────
    print(f"\n{'─'*54}")
    print("  Buscando CDI mensal (BCB série 12)…")
    print(f"{'─'*54}")
    cdi = fetch_cdi(start_date)

    return {
        "generated_at":    datetime.now().isoformat(timespec="seconds"),
        "start_date":      start_date.strftime("%Y-%m"),
        "tickers_found":   sorted(prices.keys()),
        "prices":          prices,
        "prices_daily":    prices_daily,
        "dy":              {k: v for k, v in dy.items() if v is not None},
        "dy_source":       "fundamentus" if use_fundamentus else "brapi",
        "price_source":    "yfinance_total_return" if use_yfinance else "brapi_price_only",
        "weights_history": weights_history,
        "cdi":             cdi,
    }


# ─── Refresh de dados (manual e automático) ───────────────

def _do_refresh() -> bool:
    """Executa o fetch de dados e salva o JSON. Retorna True se bem-sucedido."""
    global _refresh_state
    p = _run_params
    if not p:
        return False

    _refresh_state["running"] = True
    _refresh_state["error"]   = None
    try:
        start_date     = p["start_date"]
        token          = p.get("token")
        use_fundamentus = p.get("use_fundamentus", True)
        use_yfinance   = p.get("use_yfinance", True)

        print(f"\n  🔄  Atualizando dados ({datetime.now().strftime('%d/%m/%Y %H:%M')})…")
        data = build_dashboard_json(
            start_date,
            token=token,
            use_fundamentus=use_fundamentus,
            use_yfinance=use_yfinance,
        )
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        now_iso = datetime.now().isoformat(timespec="seconds")
        _refresh_state["last_ok"] = now_iso
        _refresh_state["error"]   = None
        print(f"  ✅  Dados atualizados: {OUTPUT_JSON.name} ({now_iso})")
        return True
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        _refresh_state["error"] = msg
        print(f"  ✗   Erro no refresh: {msg}")
        return False
    finally:
        _refresh_state["running"] = False


def _refresh_in_background():
    """Dispara _do_refresh em thread separada (evita bloquear o servidor HTTP)."""
    if _refresh_state["running"]:
        return  # já rodando
    t = threading.Thread(target=_do_refresh, daemon=True)
    t.start()


def _auto_refresh_loop(hour_brt: int = 7):
    """
    Loop em thread de fundo: dispara refresh todos os dias quando o relógio
    de Brasília (UTC-3) bater a hora configurada (default: 07:00).
    Também atualiza imediatamente se o JSON tiver mais de 23h de idade.
    """
    import time as _time

    print(f"  🕖  Auto-refresh agendado para {hour_brt:02d}:00 (horário de Brasília)")

    def _json_age_hours() -> float:
        if not OUTPUT_JSON.exists():
            return 999.0
        mtime = OUTPUT_JSON.stat().st_mtime
        return (datetime.now().timestamp() - mtime) / 3600

    # Atualiza imediatamente se o JSON tiver mais de 23h
    if _json_age_hours() > 23:
        print("  ℹ   JSON com mais de 23h — atualizando agora…")
        _do_refresh()

    last_refresh_day = None
    while True:
        _time.sleep(60)  # verifica a cada minuto
        try:
            # UTC-3 = Brasil (horário padrão de Brasília / BRT)
            now_brt = datetime.utcnow().replace(tzinfo=None)
            # Offset manual: utc - 3h
            from datetime import timedelta
            now_brt = datetime.utcnow() - timedelta(hours=3)
            today   = now_brt.date()

            # Dispara uma vez por dia na hora configurada
            if now_brt.hour == hour_brt and today != last_refresh_day:
                last_refresh_day = today
                _refresh_in_background()
        except Exception:
            pass


# ─── HTTP Server ──────────────────────────────────────────

class QuietHandler(http.server.SimpleHTTPRequestHandler):
    """Handler HTTP simples que suprime logs do servidor."""
    def log_message(self, *_):
        pass

    def log_request(self, *_):
        pass

    def _json_response(self, data: dict, status: int = 200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        """Intercepta /lamina/<portfolio> para gerar o PDF on-demand."""
        path = self.path.split('?')[0].rstrip('/')
        if path == '' or path == '/':
            self.send_response(302)
            self.send_header('Location', '/dashboard_carteiras_v2.html')
            self.end_headers()
            return

        # ── API: iniciar refresh manual ─────────────────────────
        if path == '/api/refresh':
            if _refresh_state["running"]:
                self._json_response({"status": "running", "message": "Atualização já em andamento."})
            else:
                _refresh_in_background()
                self._json_response({"status": "started", "message": "Atualização iniciada."})
            return

        # ── API: status do refresh ──────────────────────────────
        if path == '/api/status':
            age_h = None
            if OUTPUT_JSON.exists():
                age_h = round((datetime.now().timestamp() - OUTPUT_JSON.stat().st_mtime) / 3600, 1)
            self._json_response({
                "running":  _refresh_state["running"],
                "last_ok":  _refresh_state["last_ok"],
                "error":    _refresh_state["error"],
                "json_age_hours": age_h,
            })
            return

        if path.startswith('/lamina/'):
            portfolio_key = path.split('/')[-1]   # 'acoes' ou 'dividendos'
            if portfolio_key not in ('acoes', 'dividendos'):
                self.send_error(400, 'Portfolio inválido. Use /lamina/acoes ou /lamina/dividendos')
                return
            try:
                # Carrega JSON local
                json_path = SCRIPT_DIR / 'dashboard_data.json'
                if not json_path.exists():
                    self.send_error(503, 'dashboard_data.json não encontrado. Execute serve.py primeiro.')
                    return
                with open(json_path, encoding='utf-8') as f:
                    data = json.load(f)

                # Gera PDF em memória
                from gerar_factsheet import generate_pdf_bytes
                pdf_bytes = generate_pdf_bytes(portfolio_key, data)

                # Responde com o PDF
                nome_arquivo = f"factsheet_{portfolio_key}_{datetime.now().strftime('%Y-%m')}.pdf"
                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                self.send_header('Content-Disposition',
                                 f'attachment; filename="{nome_arquivo}"')
                self.send_header('Content-Length', str(len(pdf_bytes)))
                self.end_headers()
                self.wfile.write(pdf_bytes)
                print(f"  🖨   Factsheet gerado: {nome_arquivo}")
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                print(f"  ✗  Erro ao gerar factsheet: {e}\n{tb}")
                # Envia texto puro (não HTML) para o alert do browser mostrar a causa real
                body = f"Erro interno: {type(e).__name__}: {e}\n\n{tb}".encode('utf-8', errors='replace')
                self.send_response(500)
                self.send_header('Content-Type', 'text/plain; charset=utf-8')
                self.send_header('Content-Length', str(len(body)))
                self.end_headers()
                self.wfile.write(body)
        else:
            super().do_GET()


def start_server(port: int, directory: Path, open_browser: bool, html_file: str):
    os.chdir(directory)
    try:
        server = http.server.HTTPServer(("", port), QuietHandler)
    except OSError:
        print(f"\n  ⚠  Porta {port} já em uso. Tente: python3 serve.py --porta {port + 1}")
        sys.exit(1)

    url = f"http://localhost:{port}/{html_file}"
    print(f"\n{'═'*54}")
    print(f"  🌐  Dashboard:  {url}")
    print(f"  📁  Servindo:   {directory}")
    print(f"  ⏹   Parar:      Ctrl+C")
    print(f"{'═'*54}\n")

    if open_browser:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  ⏹  Servidor encerrado.")
        server.shutdown()


# ─── CLI ─────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Busca dados e serve o dashboard localmente",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument("--porta",           type=int,  default=DEFAULT_PORT, help=f"Porta HTTP (default: {DEFAULT_PORT})")
    p.add_argument("--inicio",          default="2024-06",               help="Data inicial YYYY-MM (default: 2024-06)")
    p.add_argument("--token",           default=None,                    help="Token brapi.dev (grátis em brapi.dev)")
    p.add_argument("--sem-browser",     action="store_true",             help="Não abre o browser automaticamente")
    p.add_argument("--so-servir",       action="store_true",             help="Só inicia o servidor (sem refetch)")
    p.add_argument("--sem-fundamentus", action="store_true",             help="Pula Fundamentus; usa só brapi.dev para DY")
    p.add_argument("--sem-yfinance",    action="store_true",             help="Usa brapi.dev para preços (sem retorno total por dividendos)")
    p.add_argument("--html",            default="dashboard_carteiras_v2.html", help="Arquivo HTML a abrir")
    p.add_argument("--exportar-pdf",    action="store_true",             help="Gera factsheet PDF após atualizar dados")
    p.add_argument("--mes-pdf",         default=None,                    help="Mês do factsheet YYYY-MM (default: mês atual)")
    p.add_argument("--so-pdf",          action="store_true",             help="Só gera o PDF, sem iniciar servidor")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    # Na nuvem: nunca abre browser, pula fundamentus (pode ter bloqueio de IP)
    if IS_CLOUD:
        args.sem_browser = True
        args.sem_fundamentus = True
        print("═" * 54)
        print("  CARTEIRAS TEMÁTICAS — MODO NUVEM (Railway)")
        print(f"  PORT: {DEFAULT_PORT}")
        print("═" * 54)
    else:
        print("═" * 54)
        print("  CARTEIRAS TEMÁTICAS — SERVIDOR LOCAL")
        print("═" * 54)

    if not args.so_servir:
        try:
            start_date = datetime.strptime(args.inicio, "%Y-%m").replace(day=1)
        except ValueError:
            print(f"  ✗  Formato de data inválido: '{args.inicio}'. Use YYYY-MM")
            sys.exit(1)

        # ── Fetch data ──
        try:
            data = build_dashboard_json(
                start_date,
                token=args.token,
                use_fundamentus=not args.sem_fundamentus,
                use_yfinance=not args.sem_yfinance,
            )
        except Exception as e:
            print(f"\n  ✗  Erro ao buscar dados: {e}")
            print("     Verifique sua conexão e tente novamente.")
            print("     Para iniciar só o servidor: python3 serve.py --so-servir")
            sys.exit(1)

        # ── Save JSON ──
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        n = len(data["tickers_found"])
        print(f"\n  ✅  {n} ativos salvos → {OUTPUT_JSON.name}")
    else:
        if OUTPUT_JSON.exists():
            print(f"  ℹ   Usando dados existentes: {OUTPUT_JSON.name}")
        else:
            print("  ⚠   dashboard_data.json não encontrado. Execute sem --so-servir primeiro.")

    # ── Exportar PDF ──
    if args.exportar_pdf or args.so_pdf:
        mes_pdf = args.mes_pdf or datetime.now().strftime("%Y-%m")
        factsheet_script = SCRIPT_DIR / "gerar_factsheet.py"
        if factsheet_script.exists():
            print(f"\n{'─'*54}")
            print(f"  🖨   Gerando factsheet PDF — {mes_pdf}…")
            print(f"{'─'*54}")
            cmd = [sys.executable, str(factsheet_script), "--mes", mes_pdf]
            result = subprocess.run(cmd, cwd=str(SCRIPT_DIR))
            if result.returncode == 0:
                pdf_out = SCRIPT_DIR / f"factsheet_{mes_pdf}.pdf"
                print(f"\n  ✅  Factsheet pronto: {pdf_out.name}")
            else:
                print(f"\n  ⚠  Erro ao gerar o factsheet PDF.")
        else:
            print(f"\n  ⚠  gerar_factsheet.py não encontrado em {SCRIPT_DIR}")

    if args.so_pdf:
        sys.exit(0)

    # ── Armazena parâmetros para o auto-refresh ──
    try:
        _start_date = datetime.strptime(args.inicio, "%Y-%m").replace(day=1)
    except ValueError:
        _start_date = datetime(2024, 6, 1)

    _run_params.update({
        "start_date":      _start_date,
        "token":           args.token or BRAPI_TOKEN,
        "use_fundamentus": not args.sem_fundamentus,
        "use_yfinance":    not getattr(args, "sem_yfinance", False),
    })

    # ── Inicia thread de auto-refresh diário (11:00 BRT) ──
    _ar = threading.Thread(target=_auto_refresh_loop, kwargs={"hour_brt": 11}, daemon=True)
    _ar.start()

    # ── Serve ──
    start_server(
        port=args.porta,
        directory=SCRIPT_DIR,
        open_browser=not args.sem_browser,
        html_file=args.html,
    )

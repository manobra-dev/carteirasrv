#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════
  CARTEIRAS TEMÁTICAS — GERADOR DE FACTSHEET PDF
═══════════════════════════════════════════════════════════

Uso:
    python3 gerar_factsheet.py                        → gera para o mês atual
    python3 gerar_factsheet.py --mes 2026-02          → mês específico
    python3 gerar_factsheet.py --comentario "Texto"  → comentário inline
    python3 gerar_factsheet.py --valor 100000         → valor total da carteira

Lê:
    dashboard_data.json           → preços e retornos
    rebalanceamentos_template.xlsx → pesos históricos por mês
    comentario.txt                 → texto do comentário do mês (opcional)

Gera:
    factsheet_YYYY-MM.pdf         → documento para enviar aos clientes
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

# Garante backend não-interativo ANTES de qualquer import do matplotlib.
# Essencial para funcionar dentro do serve.py (sem display/GUI).
os.environ.setdefault('MPLBACKEND', 'Agg')

SCRIPT_DIR = Path(__file__).parent

# ── Logo da empresa ─────────────────────────────────────
# Salve o arquivo PNG do logo (fundo preto, logo branco) na mesma pasta do script.
# O código remove o fundo preto automaticamente.
LOGO_PDF_PATH = SCRIPT_DIR / 'logo_eleva_branco.png'   # logo branco → headers coloridos do PDF


def _load_logo_png(path, bg_threshold: int = 55):
    """
    Carrega PNG do logo, converte fundo escuro (preto) em transparente.
    Retorna bytes PNG com canal alpha, ou None se o arquivo não existir.
    """
    try:
        from PIL import Image
        import io as _io
        img = Image.open(str(path)).convert("RGBA")
        try:
            import numpy as _np
            arr = _np.array(img, dtype=_np.uint8)
            mask = (arr[:, :, 0] < bg_threshold) & \
                   (arr[:, :, 1] < bg_threshold) & \
                   (arr[:, :, 2] < bg_threshold)
            arr[mask, 3] = 0
            img = Image.fromarray(arr)
        except ImportError:
            # Fallback sem numpy — mais lento mas funciona
            data = list(img.getdata())
            img.putdata([
                (0, 0, 0, 0)
                if r < bg_threshold and g < bg_threshold and b < bg_threshold
                else (r, g, b, a)
                for r, g, b, a in data
            ])
        buf = _io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf.read()
    except Exception:
        return None


# ── Fontes profissionais ────────────────────────────────
# Nomes internos das fontes — podem ser Lato (TTF) ou Helvetica (fallback)
_FONT_REGULAR    = 'Lato'
_FONT_BOLD       = 'Lato-Bold'
_FONT_ITALIC     = 'Lato-Italic'
_FONT_BOLDITALIC = 'Lato-BoldItalic'

_fonts_registered = False

# Arquivos Lato necessários
_LATO_FILES = {
    'Lato':            'Lato-Regular.ttf',
    'Lato-Bold':       'Lato-Bold.ttf',
    'Lato-Italic':     'Lato-Italic.ttf',
    'Lato-BoldItalic': 'Lato-BoldItalic.ttf',
}
# URLs diretas (Google Fonts / GitHub) para download automático
_LATO_URLS = {
    'Lato-Regular.ttf':    'https://raw.githubusercontent.com/google/fonts/main/ofl/lato/Lato-Regular.ttf',
    'Lato-Bold.ttf':       'https://raw.githubusercontent.com/google/fonts/main/ofl/lato/Lato-Bold.ttf',
    'Lato-Italic.ttf':     'https://raw.githubusercontent.com/google/fonts/main/ofl/lato/Lato-Italic.ttf',
    'Lato-BoldItalic.ttf': 'https://raw.githubusercontent.com/google/fonts/main/ofl/lato/Lato-BoldItalic.ttf',
}


def _find_lato_paths() -> dict:
    """
    Procura arquivos Lato em caminhos do sistema (Windows e Linux).
    Retorna dict {font_name: caminho_absoluto} para os que encontrar.
    """
    import platform
    from pathlib import Path as _P

    search_dirs = [
        SCRIPT_DIR / 'fonts',          # cache local (prioridade máxima)
        SCRIPT_DIR,                    # arquivos soltos na pasta do script
    ]
    system = platform.system()
    if system == 'Windows':
        import os
        win_root = _P(os.environ.get('WINDIR', r'C:\Windows'))
        search_dirs += [
            win_root / 'Fonts',
            _P(os.environ.get('LOCALAPPDATA', '')) / 'Microsoft' / 'Windows' / 'Fonts',
        ]
    else:
        search_dirs += [
            _P('/usr/share/fonts/truetype/lato'),
            _P('/usr/local/share/fonts/lato'),
            _P('/usr/share/fonts/lato'),
        ]

    found = {}
    for font_name, filename in _LATO_FILES.items():
        for d in search_dirs:
            p = d / filename
            try:
                if p.exists():
                    found[font_name] = str(p)
                    break
            except Exception:
                pass
    return found


def _download_lato() -> dict:
    """
    Tenta baixar os arquivos Lato para SCRIPT_DIR/fonts/.
    Retorna dict {font_name: caminho} para os que conseguiu baixar.
    """
    import urllib.request
    from pathlib import Path as _P

    cache = SCRIPT_DIR / 'fonts'
    try:
        cache.mkdir(exist_ok=True)
    except Exception:
        return {}

    downloaded = {}
    for font_name, filename in _LATO_FILES.items():
        dest = cache / filename
        if not dest.exists():
            url = _LATO_URLS[filename]
            try:
                urllib.request.urlretrieve(url, str(dest))
            except Exception:
                continue
        if dest.exists():
            downloaded[font_name] = str(dest)
    return downloaded


def _register_helvetica_alias():
    """
    Último recurso: registra 'Lato*' como aliases dos Helvetica built-in.
    Sem TTF real, mas o PDF não vai quebrar.
    """
    try:
        from reportlab.pdfbase import pdfmetrics
        mapping = [
            ('Lato',            'Helvetica'),
            ('Lato-Bold',       'Helvetica-Bold'),
            ('Lato-Italic',     'Helvetica-Oblique'),
            ('Lato-BoldItalic', 'Helvetica-BoldOblique'),
        ]
        for lato_name, helv_name in mapping:
            if lato_name not in pdfmetrics._fonts:
                pdfmetrics._fonts[lato_name] = pdfmetrics.getFont(helv_name)
    except Exception:
        pass


def _register_fonts():
    """
    Registra Lato no ReportLab (TTF) e configura matplotlib para usar Lato.
    Estratégia:
      1. Busca Lato nos caminhos do sistema (Windows e Linux).
      2. Se não encontrar, tenta baixar de raw.githubusercontent.com (cache local).
      3. Último recurso: alias Lato → Helvetica para não quebrar o PDF.
    """
    global _fonts_registered
    if _fonts_registered:
        return
    _fonts_registered = True  # marca antes para evitar reentrada

    # ── 1. Encontrar fontes no sistema ───────────────────────
    resolved = _find_lato_paths()

    # ── 2. Baixar se não tiver todas ─────────────────────────
    if len(resolved) < len(_LATO_FILES):
        downloaded = _download_lato()
        for k, v in downloaded.items():
            if k not in resolved:
                resolved[k] = v

    # ── 3. Registrar no ReportLab ────────────────────────────
    if len(resolved) == len(_LATO_FILES):
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfbase.pdfmetrics import registerFontFamily
            for name, path in resolved.items():
                pdfmetrics.registerFont(TTFont(name, path))
            registerFontFamily('Lato',
                               normal='Lato', bold='Lato-Bold',
                               italic='Lato-Italic', boldItalic='Lato-BoldItalic')
        except Exception:
            _register_helvetica_alias()
    else:
        # Fonte não encontrada e download falhou → alias seguro
        _register_helvetica_alias()

    # ── 4. Configurar matplotlib ─────────────────────────────
    try:
        import matplotlib as mpl
        import matplotlib.font_manager as fm
        for path in resolved.values():
            try:
                fm.fontManager.addfont(path)
            except Exception:
                pass
        mpl.rcParams.update({
            'font.family':     'sans-serif',
            'font.sans-serif': ['Lato', 'Liberation Sans', 'Arial', 'DejaVu Sans'],
        })
    except Exception:
        pass


# ── Cores (RGB 0-1) ─────────────────────────────────────
AZUL_ESCURO  = (0.00, 0.00, 0.24)   # #00003c  cabeçalhos e rodapé
AZUL_MEDIO   = (0.13, 0.35, 0.62)   # #215A9E  linhas alternadas
AZUL_CLARO   = (0.87, 0.92, 0.97)   # #DDE9F7  fundo linhas pares
LARANJA      = (0.85, 0.42, 0.08)   # #D96B14  subtítulos / destaque
BRANCO       = (1.00, 1.00, 1.00)
CINZA_TEXTO  = (0.20, 0.20, 0.20)
CINZA_FUNDO  = (0.96, 0.96, 0.96)
VERDE        = (0.10, 0.55, 0.20)
VERMELHO     = (0.75, 0.08, 0.08)

# ── Metadados das ações ─────────────────────────────────
METADATA = {
    "BPAC11": {"nome": "BTG Pactual",        "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "BBAS3":  {"nome": "Banco do Brasil",    "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "ITUB4":  {"nome": "Itaú Unibanco PN",   "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "JPMC34": {"nome": "JPMorgan BDR",       "setor": "Financeiro",         "liquidez": "Média",  "risco": "Médio"},
    "CPLE3":  {"nome": "Copel ON",           "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "CPLE5":  {"nome": "Copel PNA",          "setor": "Energia Elétrica",   "liquidez": "Média",  "risco": "Médio"},
    "CPLE6":  {"nome": "Copel PNB",          "setor": "Energia Elétrica",   "liquidez": "Média",  "risco": "Médio"},
    "EMBJ3":  {"nome": "Embraer",            "setor": "Aviação",            "liquidez": "Alta",   "risco": "Alto"},
    "EMBR3":  {"nome": "Embraer",            "setor": "Aviação",            "liquidez": "Alta",   "risco": "Alto"},
    "WEGE3":  {"nome": "WEG",                "setor": "Indústria",          "liquidez": "Alta",   "risco": "Médio"},
    "GGBR4":  {"nome": "Gerdau PN",          "setor": "Siderurgia",         "liquidez": "Alta",   "risco": "Alto"},
    "VALE3":  {"nome": "Vale ON",            "setor": "Mineração",          "liquidez": "Alta",   "risco": "Alto"},
    "AURA33": {"nome": "Aura Minerals BDR",  "setor": "Mineração",          "liquidez": "Média",  "risco": "Alto"},
    "AXIA3":  {"nome": "Eletrobras ON",      "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "AXIA6":  {"nome": "Eletrobras PNB",     "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "ELET3":  {"nome": "Eletrobras ON",      "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "ELET6":  {"nome": "Eletrobras PNB",     "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "TTEN3":  {"nome": "3tentos",            "setor": "Agroindústria",      "liquidez": "Média",  "risco": "Alto"},
    "CSMG3":  {"nome": "Copasa",             "setor": "Saneamento",         "liquidez": "Média",  "risco": "Médio"},
    "IVVB11": {"nome": "iShares S&P500 ETF", "setor": "Índice Global",      "liquidez": "Alta",   "risco": "Médio"},
    "BBDC4":  {"nome": "Bradesco PN",        "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "CXSE3":  {"nome": "Caixa Seguridade",   "setor": "Seguros",            "liquidez": "Alta",   "risco": "Médio"},
    "B3SA3":  {"nome": "B3 S.A.",            "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "SANB11": {"nome": "Santander Brasil",   "setor": "Financeiro",         "liquidez": "Alta",   "risco": "Médio"},
    "NDIV11": {"nome": "ETF IDIV",           "setor": "Índice Dividendos",  "liquidez": "Alta",   "risco": "Médio"},
    "PETR4":  {"nome": "Petrobras PN",       "setor": "Petroquímica",       "liquidez": "Alta",   "risco": "Alto"},
    "CPFE3":  {"nome": "CPFL Energia",       "setor": "Energia Elétrica",   "liquidez": "Alta",   "risco": "Médio"},
    "JBSS32": {"nome": "JBS BDR",            "setor": "Alimentos",          "liquidez": "Média",  "risco": "Médio"},
    "JBBS3":  {"nome": "JBS S.A.",           "setor": "Alimentos",          "liquidez": "Alta",   "risco": "Médio"},
    "BPAN4":  {"nome": "Banco Pan PN",       "setor": "Financeiro",         "liquidez": "Média",  "risco": "Alto"},
    "SUZB3":  {"nome": "Suzano",             "setor": "Papel e Celulose",   "liquidez": "Alta",   "risco": "Alto"},
    "IGTI11": {"nome": "Iguatemi units",     "setor": "Shopping",           "liquidez": "Média",  "risco": "Médio"},
    "ARZZ3":  {"nome": "Arezzo",             "setor": "Varejo",             "liquidez": "Média",  "risco": "Alto"},
    "AZZA3":  {"nome": "Azzas 2154",         "setor": "Varejo",             "liquidez": "Média",  "risco": "Alto"},
    "SMAL11": {"nome": "ETF Small Caps",     "setor": "Índice",             "liquidez": "Alta",   "risco": "Alto"},
    "BBSE3":  {"nome": "BB Seguridade",      "setor": "Seguros",            "liquidez": "Alta",   "risco": "Médio"},
}

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março",    4: "Abril",
    5: "Maio",    6: "Junho",     7: "Julho",     8: "Agosto",
    9: "Setembro",10: "Outubro",  11: "Novembro", 12: "Dezembro",
}


# ═══════════════════════════════════════════════════════════
#  1. LEITURA DE DADOS
# ═══════════════════════════════════════════════════════════

def _ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        import openpyxl
        return openpyxl


def _ensure_reportlab():
    try:
        import reportlab
        return True
    except ImportError:
        print("  ℹ  Instalando 'reportlab'…")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "reportlab", "-q"],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("  ✓  reportlab instalado.")
        return True


def ler_rebalanceamentos(xlsx_path: Path, mes: str) -> dict[str, dict[str, dict]]:
    """
    Lê rebalanceamentos_template.xlsx.

    Estrutura esperada da planilha:
      Linhas 1-3: título, instrução, legenda (ignoradas)
      Linha 4:    Ticker | Empresa | Ticker API | Jun/2024 | Jul/2024 | ...
      Linha 5+:   dados por ativo

    Retorna:
        {
            "acoes":      {"DISPLAY_TICKER": {"peso_atual": 0.10, "peso_anterior": 0.05, "api_ticker": "EMBR3"}},
            "dividendos": {...}
        }
    'peso_anterior' é o mês imediatamente anterior com dados.
    """
    openpyxl = _ensure_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    resultado = {}
    nomes_sheets = {
        "Ações": "acoes", "Acoes": "acoes", "acoes": "acoes",
        "Dividendos": "dividendos", "dividendos": "dividendos",
    }

    for sheet_name in wb.sheetnames:
        chave = nomes_sheets.get(sheet_name)
        if chave is None:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # ── Encontra a linha de cabeçalho (a que tem colunas de mês) ──
        header_idx = None
        mes_cols   = {}  # YYYY-MM → índice coluna

        for ri, row in enumerate(rows):
            found = {}
            for ci, val in enumerate(row):
                if val is None:
                    continue
                ym = _parse_mes_col(str(val).strip())
                if ym:
                    found[ym] = ci
            if len(found) >= 2:  # linha válida de meses
                header_idx = ri
                mes_cols   = found
                break

        if not mes_cols:
            print(f"  ⚠  Sheet '{sheet_name}': nenhuma coluna de mês reconhecida.")
            continue

        # Detecta quais colunas são ticker/empresa/api (antes dos meses)
        header_row  = rows[header_idx]
        first_mes_ci = min(mes_cols.values())

        # Heurística: procura "Ticker API" na linha de cabeçalho
        ticker_col    = 0   # padrão: coluna A
        api_ticker_col = None
        for ci in range(first_mes_ci):
            if header_row[ci] and "api" in str(header_row[ci]).lower():
                api_ticker_col = ci
            if header_row[ci] and "ticker" in str(header_row[ci]).lower() and "api" not in str(header_row[ci]).lower():
                ticker_col = ci

        # ── Mês atual e anterior ──────────────────────────────
        meses_sorted = sorted(mes_cols.keys())

        # Usa o mês pedido; se vazio, recua até encontrar dados
        mes_alvo = mes
        if mes not in meses_sorted:
            mes_alvo = meses_sorted[-1]
            print(f"  ℹ  Sheet '{sheet_name}': mês {mes} não está na planilha, usando {mes_alvo}.")

        # Se o mês pedido estiver na planilha mas sem dados, recua para o anterior com dados
        data_rows = rows[header_idx + 1:]
        col_alvo = mes_cols.get(mes_alvo)

        # Verifica se há dados no mês_alvo
        tem_dados = any(
            r[col_alvo] not in (None, 0, "")
            for r in data_rows
            if r and r[ticker_col] and str(r[ticker_col]).strip().upper() not in
               ("TICKER", "TOTAL", "ATIVO", "")
            and not str(r[ticker_col]).startswith("💡")
        )
        if not tem_dados:
            # Recua até encontrar mês com dados
            idx_alvo = meses_sorted.index(mes_alvo)
            for i in range(idx_alvo - 1, -1, -1):
                cand = meses_sorted[i]
                ccol = mes_cols[cand]
                if any(r[ccol] not in (None, 0, "") for r in data_rows if r and r[ticker_col]):
                    print(f"  ℹ  Sheet '{sheet_name}': mês {mes_alvo} vazio, usando {cand}.")
                    mes_alvo  = cand
                    col_alvo  = ccol
                    break

        idx_atual = meses_sorted.index(mes_alvo)
        mes_anterior = meses_sorted[idx_atual - 1] if idx_atual > 0 else None
        col_anterior = mes_cols.get(mes_anterior) if mes_anterior else None

        # ── Coleta ativos ─────────────────────────────────────
        ativos = {}
        IGNORAR = {"TICKER", "TOTAL", "ATIVO", ""}

        for row in data_rows:
            if not row:
                continue
            ticker_raw = row[ticker_col] if len(row) > ticker_col else None
            if ticker_raw is None:
                continue
            ticker = str(ticker_raw).strip().upper()
            if not ticker or ticker in IGNORAR or ticker.startswith("💡"):
                continue

            # Ticker API (para busca de dados)
            api_ticker = ticker
            if api_ticker_col is not None and len(row) > api_ticker_col and row[api_ticker_col]:
                api_ticker = str(row[api_ticker_col]).strip().upper() or ticker

            def _get_peso(ci):
                if ci is None or ci >= len(row):
                    return 0.0
                v = row[ci]
                if v is None:
                    return 0.0
                try:
                    f = float(v)
                    # normaliza: se > 1 assume percentual (ex: 10 → 0.10)
                    return f / 100.0 if f > 1.5 else f
                except (ValueError, TypeError):
                    return 0.0

            peso_atual    = _get_peso(col_alvo)
            peso_anterior = _get_peso(col_anterior)

            if peso_atual > 0 or peso_anterior > 0:
                ativos[ticker] = {
                    "peso_atual":    round(peso_atual, 4),
                    "peso_anterior": round(peso_anterior, 4),
                    "api_ticker":    api_ticker,
                    "mes_referencia": mes_alvo,
                }

        resultado[chave] = ativos
        print(f"  ✓  {sheet_name}: {len(ativos)} ativos carregados (mês: {mes_alvo})")

    return resultado


def _parse_mes_col(val: str) -> str | None:
    """Tenta converter string de coluna para YYYY-MM. Retorna None se não reconhecer."""
    import re
    val = val.strip()
    # YYYY-MM
    m = re.match(r"^(\d{4})-(\d{2})$", val)
    if m:
        return val
    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-\d{2}$", val)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{4})$", val)
    if m:
        return f"{m.group(2)}-{m.group(1).zfill(2)}"
    # Jun/2024, jan/2025, etc.
    meses_abrev = {
        "jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
        "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12,
        "feb":2,"apr":4,"may":5,"aug":8,"sep":9,"oct":10,"dec":12,
    }
    m = re.match(r"^([a-zA-Z]{3})[/\-](\d{4})$", val)
    if m:
        num = meses_abrev.get(m.group(1).lower())
        if num:
            return f"{m.group(2)}-{str(num).zfill(2)}"
    return None


def ler_comentario(mes: str) -> str:
    """Lê comentario.txt ou retorna texto padrão."""
    path = SCRIPT_DIR / "comentario.txt"
    if path.exists():
        txt = path.read_text(encoding="utf-8").strip()
        if txt:
            return txt
    return (
        f"Neste mês, as carteiras mantiveram posicionamento consistente com o "
        f"cenário macroeconômico em vigor. O portfólio foi ajustado para capturar "
        f"oportunidades em setores com melhor relação risco/retorno. "
        f"Acompanhe as movimentações detalhadas na seção anterior."
    )


def ler_retornos_json(json_path: Path, mes: str) -> dict:
    """
    Lê dashboard_data.json e extrai retornos acumulados por ticker até o mês informado.
    O JSON pode ter estrutura flat {prices: {ticker: {YYYY-MM: price}}}
    ou aninhada {portfolios: {key: {prices: {ticker: {YYYY-MM: price}}}}}.
    Retorna {ticker: retorno_pct} — dicionário flat, consultado por qualquer portfólio.
    """
    if not json_path.exists():
        return {}

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    ticker_retornos = {}

    # ── Estrutura flat (serve.py atual) ──────────────────────
    all_prices_dicts = []
    if "prices" in data and isinstance(data["prices"], dict):
        # Verifica se é {ticker: {YYYY-MM: price}} ou {YYYY-MM: price}
        sample = next(iter(data["prices"].values()), None)
        if isinstance(sample, dict):
            all_prices_dicts.append(data["prices"])

    # ── Estrutura aninhada por portfólio ──────────────────────
    for pdata in data.get("portfolios", {}).values():
        if isinstance(pdata, dict) and "prices" in pdata:
            all_prices_dicts.append(pdata["prices"])

    for prices_dict in all_prices_dicts:
        for ticker, monthly in prices_dict.items():
            if not isinstance(monthly, dict) or ticker in ticker_retornos:
                continue
            meses = sorted(monthly.keys())
            if not meses:
                continue
            primeiro = meses[0]
            alvo = mes if mes in monthly else meses[-1]
            p0 = monthly.get(primeiro)
            p1 = monthly.get(alvo)
            if p0 and p0 > 0 and p1:
                ticker_retornos[ticker] = round((p1 / p0 - 1) * 100, 2)

    return ticker_retornos


# ═══════════════════════════════════════════════════════════
#  2. GERAÇÃO DO PDF
# ═══════════════════════════════════════════════════════════

def gerar_pdf(mes: str, comentario_txt: str, valor_total: float,
              rebal: dict, retornos: dict, output_path: Path):
    """Gera o factsheet PDF completo."""
    _ensure_reportlab()
    _register_fonts()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, KeepTogether,
    )
    from reportlab.platypus import BalancedColumns
    from reportlab.lib.colors import HexColor, Color

    # Cores ReportLab
    COR_AZUL    = Color(*AZUL_ESCURO)
    COR_AZUL_M  = Color(*AZUL_MEDIO)
    COR_AZUL_CL = Color(*AZUL_CLARO)
    COR_LARANJA = Color(*LARANJA)
    COR_BRANCO  = colors.white
    COR_VERDE   = Color(*VERDE)
    COR_VERMELHO = Color(*VERMELHO)
    COR_CINZA   = Color(0.94, 0.94, 0.94)

    # Estilos
    estilo_base = ParagraphStyle(
        "base", fontName="Lato", fontSize=9,
        textColor=colors.Color(*CINZA_TEXTO), leading=13,
    )
    estilo_titulo_pag = ParagraphStyle(
        "titulo_pag", fontName="Lato-Bold", fontSize=17,
        textColor=COR_BRANCO, leading=20, spaceAfter=0,
    )
    estilo_subtitulo = ParagraphStyle(
        "subtitulo", fontName="Lato-Bold", fontSize=11,
        textColor=COR_LARANJA, leading=14, spaceBefore=10, spaceAfter=4,
    )
    estilo_secao = ParagraphStyle(
        "secao", fontName="Lato-Bold", fontSize=10,
        textColor=COR_AZUL, leading=13, spaceBefore=8, spaceAfter=3,
    )
    estilo_corpo = ParagraphStyle(
        "corpo", fontName="Lato", fontSize=9,
        textColor=colors.Color(*CINZA_TEXTO), leading=14, alignment=TA_JUSTIFY,
    )
    estilo_rodape = ParagraphStyle(
        "rodape", fontName="Lato", fontSize=7,
        textColor=colors.Color(0.5, 0.5, 0.5), leading=10, alignment=TA_CENTER,
    )
    estilo_cel = ParagraphStyle(
        "cel", fontName="Lato", fontSize=8.5,
        textColor=colors.Color(*CINZA_TEXTO), leading=11,
    )
    estilo_cel_bold = ParagraphStyle(
        "cel_bold", fontName="Lato-Bold", fontSize=8.5,
        textColor=colors.Color(*AZUL_ESCURO), leading=11,
    )
    estilo_cel_centro = ParagraphStyle(
        "cel_centro", fontName="Lato", fontSize=8.5,
        textColor=colors.Color(*CINZA_TEXTO), leading=11, alignment=TA_CENTER,
    )
    estilo_cel_dir = ParagraphStyle(
        "cel_dir", fontName="Lato", fontSize=8.5,
        textColor=colors.Color(*CINZA_TEXTO), leading=11, alignment=TA_RIGHT,
    )

    # Dimensões
    PAGE_W, PAGE_H = A4
    M_L, M_R, M_T, M_B = 1.8*cm, 1.8*cm, 2.2*cm, 2.0*cm
    CONTENT_W = PAGE_W - M_L - M_R

    # Mês em português
    ano, mo = int(mes[:4]), int(mes[5:7])
    mes_nome = MESES_PT.get(mo, "?")
    mes_label = f"{mes_nome} {ano}"

    story = []

    # ── Função auxiliar: cabeçalho de seção ──────────────────

    def header_secao(titulo: str, subtitulo_txt: str = "") -> list:
        """Retorna flowables para o cabeçalho azul de seção."""
        elements = []
        # Barra azul escuro
        data = [[Paragraph(f'<font color="white"><b>{titulo}</b></font>',
                           ParagraphStyle("h", fontName="Lato-Bold", fontSize=13,
                                          textColor=COR_BRANCO, leading=16))]]
        t = Table(data, colWidths=[CONTENT_W])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), COR_AZUL),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ]))
        elements.append(t)
        if subtitulo_txt:
            elements.append(Spacer(1, 3))
            elements.append(Paragraph(subtitulo_txt, estilo_subtitulo))
        return elements

    # ── Função: tabela de portfólio ───────────────────────────

    def tabela_portfolio(cart_key: str, cart_nome: str) -> list:
        elements = []
        ativos = rebal.get(cart_key, {})
        ret_dict = retornos  # flat dict {ticker: retorno_pct}

        # Cabeçalho
        elements += header_secao(
            f"CARTEIRAS TEMÁTICAS — {cart_nome.upper()}",
            f"Portfólio Proposto – {mes_label}",
        )
        elements.append(Spacer(1, 6))

        if not ativos:
            elements.append(Paragraph("⚠ Dados de composição não disponíveis.", estilo_corpo))
            return elements

        # Filtra apenas ativos com peso atual > 0
        ativos_ativos = {t: v for t, v in ativos.items() if v.get("peso_atual", 0) > 0}

        # Cabeçalho da tabela
        col_headers = ["Ativo", "Nome", "Setor", "Liquidez", "Risco", "Peso", "Var. Mês", "Retorno*"]
        col_widths   = [1.65*cm, 4.2*cm, 2.9*cm, 1.55*cm, 1.7*cm, 1.35*cm, 1.5*cm, 1.65*cm]

        def _cel(txt, estilo=estilo_cel):
            return Paragraph(xml_escape(str(txt)), estilo)

        def _cor_ret(val):
            if val is None:
                return colors.Color(*CINZA_TEXTO)
            return COR_VERDE if val >= 0 else COR_VERMELHO

        header_row = [
            Paragraph(f'<font color="white"><b>{h}</b></font>',
                      ParagraphStyle("th", fontName="Lato-Bold", fontSize=8,
                                     textColor=COR_BRANCO, leading=11, alignment=TA_CENTER))
            for h in col_headers
        ]

        data_rows = [header_row]
        total_peso = 0.0

        for i, (ticker, vals) in enumerate(sorted(ativos_ativos.items())):
            meta = METADATA.get(ticker, {})
            nome   = meta.get("nome",  ticker)   # _cel() faz xml_escape
            setor  = meta.get("setor", "—")      # _cel() faz xml_escape
            liq    = meta.get("liquidez", "—")
            risco  = meta.get("risco",    "—")
            peso   = vals.get("peso_atual", 0)
            p_ant  = vals.get("peso_anterior", 0)
            var_mes = round((peso - p_ant) * 100, 1) if p_ant else 0.0
            ret    = ret_dict.get(ticker)
            total_peso += peso

            # Cor da variação mensal
            var_txt = f"{var_mes:+.1f}%" if var_mes != 0 else "—"
            ret_txt = f"{ret:+.1f}%" if ret is not None else "—"

            var_estilo = ParagraphStyle(
                f"v{i}", fontName="Lato-Bold", fontSize=8,
                textColor=COR_VERDE if var_mes >= 0 else COR_VERMELHO,
                leading=11, alignment=TA_CENTER,
            )
            ret_estilo = ParagraphStyle(
                f"r{i}", fontName="Lato-Bold", fontSize=8,
                textColor=_cor_ret(ret), leading=11, alignment=TA_CENTER,
            )

            row = [
                Paragraph(f"<b>{xml_escape(ticker)}</b>", estilo_cel_bold),
                _cel(nome),
                _cel(setor),
                _cel(liq, estilo_cel_centro),
                _cel(risco, estilo_cel_centro),
                Paragraph(f"<b>{peso*100:.1f}%</b>", ParagraphStyle(
                    f"p{i}", fontName="Lato-Bold", fontSize=8,
                    textColor=COR_AZUL, leading=11, alignment=TA_CENTER)),
                Paragraph(var_txt, var_estilo),
                Paragraph(ret_txt, ret_estilo),
            ]
            data_rows.append(row)

        # Linha TOTAL
        total_pct = f"{total_peso*100:.0f}%"  # "100%" sem decimal para caber na coluna
        total_row = [
            Paragraph("<b>TOTAL</b>", estilo_cel_bold),
            _cel(""), _cel(""), _cel(""), _cel(""),
            Paragraph(f"<b>{total_pct}</b>", ParagraphStyle(
                "tot", fontName="Lato-Bold", fontSize=8.5,
                textColor=COR_AZUL, leading=11, alignment=TA_CENTER)),
            _cel(""), _cel(""),
        ]
        data_rows.append(total_row)

        tbl = Table(data_rows, colWidths=col_widths, repeatRows=1)

        # Estilo tabela
        ts = [
            # Cabeçalho
            ("BACKGROUND",    (0, 0), (-1, 0),  COR_AZUL_M),
            ("TOPPADDING",    (0, 0), (-1, 0),  5),
            ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
            # Linhas alternadas
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [COR_BRANCO, COR_AZUL_CL]),
            # Linha total
            ("BACKGROUND",    (0, -1), (-1, -1), Color(0.87, 0.91, 0.96)),
            ("TOPPADDING",    (0, -1), (-1, -1), 5),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 5),
            # Geral
            ("TOPPADDING",    (0, 1), (-1, -2), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -2), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("GRID",          (0, 0), (-1, -1), 0.25, colors.Color(0.8, 0.85, 0.92)),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.0,  COR_AZUL),
            ("LINEBELOW",     (0, -2),(-1, -2), 1.0,  COR_AZUL_M),
        ]
        tbl.setStyle(TableStyle(ts))
        elements.append(KeepTogether(tbl))

        elements.append(Spacer(1, 4))
        elements.append(Paragraph(
            f"* Retorno acumulado desde o início do monitoramento (retorno total, incluindo dividendos reinvestidos via yfinance).",
            ParagraphStyle("nota", fontName="Lato-Italic", fontSize=7,
                           textColor=colors.Color(0.5, 0.5, 0.5), leading=10),
        ))

        return elements

    # ── Função: seção de movimentações ───────────────────────

    def tabela_movimentacoes() -> list:
        elements = []
        elements += header_secao("MOVIMENTAÇÃO DAS CARTEIRAS", f"Ajustes realizados em {mes_label}")
        elements.append(Spacer(1, 8))

        for cart_key, cart_nome in [("acoes", "Carteira de Ações"), ("dividendos", "Carteira de Dividendos")]:
            ativos = rebal.get(cart_key, {})
            entradas = []
            saidas   = []
            aumentos = []
            reducoes = []

            for ticker, vals in ativos.items():
                p_atual = vals.get("peso_atual", 0)
                p_ant   = vals.get("peso_anterior", 0)
                diff    = round((p_atual - p_ant) * 100, 1)
                if p_ant == 0 and p_atual > 0:
                    entradas.append((ticker, p_atual * 100))
                elif p_atual == 0 and p_ant > 0:
                    saidas.append((ticker, p_ant * 100))
                elif diff > 0:
                    aumentos.append((ticker, diff))
                elif diff < 0:
                    reducoes.append((ticker, diff))

            elements.append(Paragraph(f"<b>{cart_nome}</b>", estilo_secao))

            if not any([entradas, saidas, aumentos, reducoes]):
                elements.append(Paragraph("Sem alterações neste mês.", estilo_corpo))
                elements.append(Spacer(1, 6))
                continue

            # Tabela de movimentações
            mov_data = []

            def _mov_header(txt):
                return Paragraph(f'<font color="white"><b>{txt}</b></font>',
                                 ParagraphStyle("mh", fontName="Lato-Bold", fontSize=8,
                                                textColor=COR_BRANCO, leading=11, alignment=TA_CENTER))

            col_w = [CONTENT_W / 4] * 4
            mov_data.append([
                _mov_header("COMPRAS (Entrada)"),
                _mov_header("VENDAS (Saída)"),
                _mov_header("Aumento de Posição"),
                _mov_header("Redução de Posição"),
            ])

            max_rows = max(len(entradas), len(saidas), len(aumentos), len(reducoes), 1)
            for j in range(max_rows):
                def _fmt_entrada(lst, j, verde=True):
                    if j < len(lst):
                        t, v = lst[j]
                        ts = xml_escape(t)
                        return Paragraph(
                            f'<font color="#{"1A8C33" if verde else "BF1414"}"><b>{ts}</b></font> '
                            f'({v:.1f}%)',
                            ParagraphStyle(f"mj{j}", fontName="Lato", fontSize=8.5, leading=12)
                        )
                    return Paragraph("", estilo_cel)

                def _fmt_variacao(lst, j, verde=True):
                    if j < len(lst):
                        t, v = lst[j]
                        ts = xml_escape(t)
                        sinal = "+" if verde else "-"
                        cor_hex = "1A8C33" if verde else "BF1414"
                        return Paragraph(
                            f'<font color="#{cor_hex}"><b>{ts}</b> {sinal}{abs(v):.1f}pp</font>',
                            ParagraphStyle(f"mv{j}", fontName="Lato", fontSize=8.5, leading=12)
                        )
                    return Paragraph("", estilo_cel)

                row = [
                    _fmt_entrada(entradas, j, verde=True),
                    _fmt_entrada(saidas,   j, verde=False),
                    _fmt_variacao(aumentos, j, verde=True),
                    _fmt_variacao(reducoes, j, verde=False),
                ]
                mov_data.append(row)

            mov_tbl = Table(mov_data, colWidths=col_w)
            mov_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0),  COR_AZUL_M),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [COR_BRANCO, COR_AZUL_CL]),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
                ("GRID",          (0, 0), (-1, -1), 0.25, colors.Color(0.85, 0.88, 0.93)),
                ("LINEBELOW",     (0, 0), (-1, 0),  1.0,  COR_AZUL),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(KeepTogether(mov_tbl))
            elements.append(Spacer(1, 12))

        return elements

    # ── Função: comentário do gestor ─────────────────────────

    def secao_comentario(texto: str) -> list:
        elements = []
        elements += header_secao("COMENTÁRIO DO GESTOR", f"Análise de Mercado — {mes_label}")
        elements.append(Spacer(1, 10))

        # Divide o texto em parágrafos (linhas em branco)
        paragrafos = [p.strip() for p in texto.split("\n\n") if p.strip()]
        if not paragrafos:
            paragrafos = [texto.strip()]

        for par in paragrafos:
            elements.append(Paragraph(xml_escape(par), estilo_corpo))
            elements.append(Spacer(1, 8))

        # Assinatura
        elements.append(Spacer(1, 6))
        elements.append(HRFlowable(width=CONTENT_W, thickness=0.5,
                                   color=colors.Color(0.7, 0.7, 0.7)))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(
            f"<i>Tematica Investimentos  |  {mes_label}</i>",
            ParagraphStyle("ass", fontName="Lato-Italic", fontSize=8,
                           textColor=colors.Color(0.5, 0.5, 0.5), alignment=TA_RIGHT),
        ))
        return elements

    # ── Callbacks de página (cabeçalho / rodapé) ─────────────

    def _draw_page(canvas, doc):
        canvas.saveState()
        # Cabeçalho
        canvas.setFillColorRGB(*AZUL_ESCURO)
        canvas.rect(0, PAGE_H - 1.3*cm, PAGE_W, 1.3*cm, fill=1, stroke=0)
        canvas.setFillColorRGB(*LARANJA)
        canvas.rect(0, PAGE_H - 1.5*cm, PAGE_W, 0.2*cm, fill=1, stroke=0)

        canvas.setFillColorRGB(1, 1, 1)
        canvas.setFont("Lato-Bold", 11)
        canvas.drawString(M_L, PAGE_H - 0.95*cm, "CARTEIRAS TEMÁTICAS")
        canvas.setFont("Lato", 9)
        canvas.drawRightString(PAGE_W - M_R, PAGE_H - 0.95*cm, f"Factsheet {mes_label}")

        # Rodapé
        canvas.setFillColorRGB(*AZUL_ESCURO)
        canvas.rect(0, 0, PAGE_W, 1.0*cm, fill=1, stroke=0)
        canvas.setFillColorRGB(*LARANJA)
        canvas.rect(0, 1.0*cm, PAGE_W, 0.15*cm, fill=1, stroke=0)

        canvas.setFillColorRGB(1, 1, 1)
        canvas.setFont("Lato", 7)
        canvas.drawString(M_L, 0.35*cm,
            "Este material é de caráter informativo e não constitui oferta de valores mobiliários. "
            "Rentabilidade passada não é garantia de rentabilidade futura.")
        canvas.drawRightString(PAGE_W - M_R, 0.35*cm, f"Página {doc.page}")

        canvas.restoreState()

    # ── Monta o story ─────────────────────────────────────────

    # Capa / Resumo
    story += header_secao("CARTEIRAS TEMÁTICAS", f"Factsheet — {mes_label}")
    story.append(Spacer(1, 10))

    # Resumo das carteiras lado a lado
    resumo_cols = []
    for cart_key, cart_nome in [("acoes", "Carteira de Ações"), ("dividendos", "Carteira de Dividendos")]:
        ativos   = rebal.get(cart_key, {})
        n_ativos = sum(1 for v in ativos.values() if v.get("peso_atual", 0) > 0)
        # Retorno médio ponderado (retornos é flat {ticker: pct})
        ret_pond = 0.0
        for ticker, vals in ativos.items():
            peso = vals.get("peso_atual", 0)
            ret  = retornos.get(ticker, 0) or 0
            ret_pond += peso * ret

        bloco = [
            Paragraph(f"<b>{cart_nome}</b>", estilo_secao),
            Paragraph(f"Número de ativos: <b>{n_ativos}</b>", estilo_cel),
            Paragraph(f"Retorno ponderado acum.: <b>{ret_pond:+.1f}%</b>", estilo_cel),
            Paragraph(f"Mês de referência: <b>{mes_label}</b>", estilo_cel),
        ]
        resumo_cols.append(bloco)

    if resumo_cols:
        half_w = CONTENT_W / 2 - 0.5*cm
        for bloco in resumo_cols:
            for el in bloco:
                story.append(el)
            story.append(Spacer(1, 4))

    story.append(Spacer(1, 10))
    story.append(PageBreak())

    # Pág 2: Carteira de Ações
    story += tabela_portfolio("acoes", "Carteira de Ações")
    story.append(Spacer(1, 12))
    story.append(PageBreak())

    # Pág 3: Carteira de Dividendos
    story += tabela_portfolio("dividendos", "Carteira de Dividendos")
    story.append(Spacer(1, 12))
    story.append(PageBreak())

    # Pág 4: Movimentações
    story += tabela_movimentacoes()
    story.append(PageBreak())

    # Pág 5: Comentário
    story += secao_comentario(comentario_txt)

    # ── Build ─────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=M_L, rightMargin=M_R,
        topMargin=M_T + 1.5*cm, bottomMargin=M_B + 1.0*cm,
        title=f"Factsheet Carteiras Temáticas {mes_label}",
        author="Tematica Investimentos",
        subject="Carteiras Temáticas — Renda Variável",
    )
    doc.build(story, onFirstPage=_draw_page, onLaterPages=_draw_page)
    print(f"  ✓  PDF gerado: {output_path}")


# ═══════════════════════════════════════════════════════════
#  3. MAIN
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Gera factsheet PDF das Carteiras Temáticas")
    parser.add_argument("--mes",         default=None,
                        help="Mês de referência YYYY-MM (padrão: mês atual)")
    parser.add_argument("--comentario",  default=None,
                        help="Texto do comentário do gestor (inline)")
    parser.add_argument("--valor",       type=float, default=100_000.0,
                        help="Valor total da carteira em R$ (padrão: 100.000)")
    parser.add_argument("--excel",       default=None,
                        help="Caminho do Excel de rebalanceamentos (padrão: rebalanceamentos_template.xlsx)")
    parser.add_argument("--json",        default=None,
                        help="Caminho do dashboard_data.json (padrão: dashboard_data.json)")
    parser.add_argument("--saida",       default=None,
                        help="Arquivo de saída (padrão: factsheet_YYYY-MM.pdf)")
    args = parser.parse_args()

    # Mês — padrão: mês anterior (último mês fechado)
    if args.mes:
        mes = args.mes
    else:
        _now = datetime.now()
        _m = _now.month - 1 or 12
        _y = _now.year if _now.month > 1 else _now.year - 1
        mes = f"{_y}-{_m:02d}"

    print(f"\n{'═'*55}")
    print(f"  GERADOR DE FACTSHEET — {mes}")
    print(f"{'═'*55}\n")

    # Caminhos
    xlsx_path = Path(args.excel) if args.excel else SCRIPT_DIR / "rebalanceamentos_template.xlsx"
    json_path = Path(args.json)  if args.json  else SCRIPT_DIR / "dashboard_data.json"
    if args.saida:
        output_path = Path(args.saida)
    else:
        output_path = SCRIPT_DIR / f"factsheet_{mes}.pdf"

    # 1. Rebalanceamentos
    print("  📊  Lendo rebalanceamentos…")
    if xlsx_path.exists():
        rebal = ler_rebalanceamentos(xlsx_path, mes)
        for k, v in rebal.items():
            print(f"    ✓  {k}: {len(v)} ativos carregados")
    else:
        print(f"  ⚠  {xlsx_path} não encontrado. Usando dados padrão do serve.py…")
        # Fallback: lê do JSON
        rebal = {}

    # 2. Retornos do JSON
    print("  📈  Lendo retornos do dashboard_data.json…")
    if json_path.exists():
        retornos = ler_retornos_json(json_path, mes)
        print(f"    ✓  {len(retornos)} retornos de tickers carregados")
    else:
        print(f"  ⚠  {json_path} não encontrado. Retornos não disponíveis.")
        retornos = {}

    # 3. Comentário
    if args.comentario:
        comentario_txt = args.comentario
    else:
        comentario_txt = ler_comentario(mes)
    print(f"  📝  Comentário: {comentario_txt[:60]}…" if len(comentario_txt) > 60
          else f"  📝  Comentário carregado ({len(comentario_txt)} chars)")

    # 4. Gera PDF
    print(f"\n  🖨   Gerando PDF…")
    gerar_pdf(mes, comentario_txt, args.valor, rebal, retornos, output_path)

    print(f"\n{'═'*55}")
    print(f"  ✅  Factsheet salvo em: {output_path.name}")
    print(f"{'═'*55}\n")


# ═══════════════════════════════════════════════════════════
#  4. FACTSHEET POR CARTEIRA — chamado pelo serve.py via /lamina/
#     Gera PDF individual (2 páginas) para cada carteira com
#     gráfico de performance (matplotlib) e tabela mensal.
# ═══════════════════════════════════════════════════════════

_PORTFOLIO_CFG = {
    'acoes': {
        'name':        'Carteira de Ações',
        'style':       'Valor / Growth',
        'bench':       'BOVA11',
        'bench_label': 'IBOVESPA (BOVA11)',
        'bench_short': 'IBOV',
        'color':       '#00003c',
        'color_mid':   '#1a1a7e',
        'color_light': '#ebf8ff',
        'tickers': ['BPAC11','BBAS3','ITUB4','JPMC34','CPLE3','EMBJ3',
                    'WEGE3','GGBR4','VALE3','AURA33','AXIA3','TTEN3','CSMG3','IVVB11'],
        'weights': [5,5,12,5,8,10,5,5,10,10,5,6,4,10],
        'sectors': ['Financeiro','Financeiro','Financeiro','Financeiro',
                    'Energia Elétrica','Aviação','Indústria','Siderurgia',
                    'Mineração','Mineração','Energia Elétrica','Agroindústria',
                    'Saneamento','Índices'],
    },
    'dividendos': {
        'name':        'Carteira de Dividendos',
        'style':       'Yield / Renda',
        'bench':       'DIVO11',
        'bench_label': 'IDIV (DIVO11)',
        'bench_short': 'IDIV',
        'color':       '#276749',
        'color_mid':   '#38a169',
        'color_light': '#f0fff4',
        'tickers': ['BBDC4','CXSE3','ITUB4','B3SA3','SANB11',
                    'NDIV11','PETR4','VALE3','CSMG3','AXIA6'],
        'weights': [10,8,15,5,5,10,12,10,15,10],
        'sectors': ['Financeiro','Seguros','Financeiro','Financeiro','Financeiro',
                    'Índice','Petróleo & Gás','Mineração','Saneamento','Energia Elétrica'],
    },
}

_SECTOR_COLORS_HEX = {
    'Financeiro':       '#2b6cb0',
    'Energia Elétrica': '#d69e2e',
    'Aviação':          '#553c9a',
    'Indústria':        '#4a5568',
    'Siderurgia':       '#744210',
    'Mineração':        '#276749',
    'Agroindústria':    '#22543d',
    'Saneamento':       '#2c7a7b',
    'Índices':          '#718096',
    'Índice':           '#718096',
    'Seguros':          '#9f7aea',
    'Petróleo & Gás':   '#975a16',
}

_MESES_ABREV = ['Jan','Fev','Mar','Abr','Mai','Jun',
                 'Jul','Ago','Set','Out','Nov','Dez']


def _ml(ym: str) -> str:
    """'2026-02' → 'Fev/26'"""
    try:
        y, m = ym.split('-')
        return f"{_MESES_ABREV[int(m)-1]}/{y[2:]}"
    except Exception:
        return ym


def _compute_monthly_cart(prices: dict, cfg: dict, wh: dict) -> dict:
    """
    Calcula retorno mensal ponderado da carteira.
    Algoritmo idêntico ao dashboard JavaScript:
    - Usa o elemento ANTERIOR no array do próprio ticker (idx-1), não um prev_m global
    - Se há weights_history para o mês usa esses pesos (0 para tickers ausentes)
    - Se não há history, usa pesos estáticos
    - Não normaliza pelo peso total (mesma lógica do JS: retorna wRet, não wRet/wSum)
    """
    # Pré-computa lista e índice de meses por ticker — busca O(1)
    ticker_sorted  = {}   # ticker → lista de meses ordenada
    ticker_idx_map = {}   # ticker → {mes: índice}
    for t in cfg['tickers']:
        if t in prices:
            ms = sorted(prices[t].keys())
            ticker_sorted[t]  = ms
            ticker_idx_map[t] = {m: i for i, m in enumerate(ms)}

    all_months = sorted(set(m for ms in ticker_sorted.values() for m in ms))

    result = {}
    for curr_m in all_months:
        has_history = curr_m in wh
        wr, ws = 0.0, 0.0
        for i, t in enumerate(cfg['tickers']):
            if t not in ticker_idx_map:
                continue
            idx2 = ticker_idx_map[t].get(curr_m, -1)
            if idx2 <= 0:          # sem dado nesse mês ou primeiro ponto (sem prev)
                continue
            p0 = float(prices[t][ticker_sorted[t][idx2 - 1]])
            p1 = float(prices[t][curr_m])
            if not p0:
                continue
            ret = (p1 / p0) - 1
            # Mesma lógica do dashboard:
            # history presente → usa peso do histórico (0 se ticker ausente no mês)
            # sem history → usa peso estático
            if has_history:
                w = wh[curr_m].get(t, 0) / 100
            else:
                w = cfg['weights'][i] / 100
            if w == 0:
                continue
            wr += w * ret
            ws += w
        if ws >= 0.5:
            result[curr_m] = wr * 100   # percentual, sem normalizar — igual ao dashboard JS
    return result


def _compute_monthly_bench(prices: dict, bench: str) -> dict:
    pm = prices.get(bench, {})
    months = sorted(pm.keys())
    result = {}
    for i in range(1, len(months)):
        p0, p1 = pm.get(months[i-1]), pm.get(months[i])
        if p0 and p1 and float(p0) != 0:
            result[months[i]] = (float(p1) / float(p0) - 1) * 100
    return result


def _cumulative(monthly: dict) -> dict:
    cum, result = 1.0, {}
    for m in sorted(monthly):
        cum *= 1 + monthly[m] / 100
        result[m] = (cum - 1) * 100
    return result


def _fmt_pct(v, dec=2) -> str:
    if v is None:
        return '—'
    s = '+' if v >= 0 else ''
    return f"{s}{v:.{dec}f}%".replace('.', ',')


def _chart_perf_png(cart_m: dict, bench_m: dict, cfg: dict) -> bytes:
    """Gráfico de linha acumulado — retorna PNG bytes."""
    import io as _io
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    all_m = sorted(set(cart_m) | set(bench_m))
    if not all_m:
        return b''
    cc = _cumulative(cart_m)
    bc = _cumulative(bench_m)

    cx = [i for i, m in enumerate(all_m) if cc.get(m) is not None]
    cy = [cc[all_m[i]] for i in cx]
    bx = [i for i, m in enumerate(all_m) if bc.get(m) is not None]
    by = [bc[all_m[i]] for i in bx]

    fig, ax = plt.subplots(figsize=(6.6, 2.7), dpi=155)
    ax.set_facecolor('#fafafa')
    fig.patch.set_facecolor('white')

    if cy:
        ax.fill_between(cx, cy, 0, alpha=0.09, color=cfg['color'])
    ax.axhline(0, color='#cbd5e0', linewidth=0.6, zorder=0)
    if cy:
        ax.plot(cx, cy, color=cfg['color'],   lw=1.6, label=cfg['name'],
                solid_capstyle='round', zorder=3)
    if by:
        ax.plot(bx, by, color='#a0aec0', lw=1.1, ls='--',
                label=cfg['bench_short'], zorder=2)

    step = max(1, len(all_m) // 14)
    ticks = list(range(0, len(all_m), step))
    ax.set_xticks(ticks)
    ax.set_xticklabels([_ml(all_m[i]) for i in ticks],
                       fontsize=6, rotation=35, ha='right')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:+.1f}%'))
    ax.tick_params(axis='y', labelsize=6)
    for sp in ['top', 'right']:
        ax.spines[sp].set_visible(False)
    for sp in ['left', 'bottom']:
        ax.spines[sp].set_color('#e2e8f0')
    ax.grid(axis='y', color='#e2e8f0', lw=0.5, alpha=0.8, zorder=0)
    ax.set_xlim(-0.3, len(all_m) - 0.7)
    ax.legend(loc='upper left', fontsize=6.5, framealpha=0.85,
              edgecolor='#e2e8f0', fancybox=False)
    plt.tight_layout(pad=0.4)

    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=155, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _chart_sector_png(cfg: dict) -> bytes:
    """Gráfico de pizza por setor — retorna PNG bytes."""
    import io as _io
    import matplotlib.pyplot as plt

    sm = {}
    for t, w, s in zip(cfg['tickers'], cfg['weights'], cfg['sectors']):
        sm[s] = sm.get(s, 0) + w
    labels = list(sm.keys())
    sizes  = list(sm.values())
    clrs   = [_SECTOR_COLORS_HEX.get(l, '#718096') for l in labels]

    fig, ax = plt.subplots(figsize=(3.1, 3.1), dpi=150)
    _, _, autotexts = ax.pie(
        sizes, colors=clrs, startangle=90,
        autopct=lambda p: f'{p:.0f}%' if p >= 6 else '',
        pctdistance=0.70,
        wedgeprops=dict(linewidth=1.8, edgecolor='white'),
    )
    for at in autotexts:
        at.set_fontsize(6.5)
        at.set_color('white')
        at.set_fontweight('bold')
    ax.set_aspect('equal')
    fig.patch.set_facecolor('white')
    plt.tight_layout(pad=0.1)

    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _risk_periods(monthly: dict, cur_year: str):
    """
    Retorna (periods_labels, vol_values, dd_values) para 4 períodos.
    monthly: {YYYY-MM: retorno_%}
    """
    import math, statistics

    months_sorted = sorted(monthly.keys())
    rets_all = [monthly[m] / 100 for m in months_sorted]  # decimais

    def _slice(period):
        if period == 'all':
            return rets_all
        if period == 'ytd':
            return [monthly[m] / 100 for m in months_sorted if m.startswith(cur_year)]
        if period == '12m':
            return rets_all[-12:]
        if period == '3m':
            return rets_all[-3:]
        return rets_all

    def _vol(r):
        if len(r) < 2:
            return None
        return statistics.stdev(r) * math.sqrt(12) * 100

    def _maxdd(r):
        if not r:
            return None
        cum, peak, dd = 1.0, 1.0, 0.0
        for ret in r:
            cum *= (1 + ret)
            if cum > peak:
                peak = cum
            d = (cum - peak) / peak
            if d < dd:
                dd = d
        return dd * 100  # negativo

    periods = ['3m', 'ytd', '12m', 'all']
    labels  = ['Últ. 3M', 'YTD', '12M', 'Desde Início']
    vols = [_vol(_slice(p))  for p in periods]
    dds  = [_maxdd(_slice(p)) for p in periods]
    return labels, vols, dds


def _chart_vol_dd_png(cart_m: dict, bench_m: dict, cfg: dict) -> bytes:
    """Gráfico duplo: Volatilidade Anualizada e Máximo Drawdown por período."""
    import io as _io
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np
    from datetime import datetime as _dt

    cur_year = str(_dt.now().year)
    p_labels, vol_c, dd_c = _risk_periods(cart_m,  cur_year)
    _,        vol_b, dd_b = _risk_periods(bench_m, cur_year)

    port_col  = cfg['color']
    bench_col = '#a0aec0'
    x = np.arange(len(p_labels))
    bw = 0.35

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(7.2, 2.8), dpi=150)
    fig.patch.set_facecolor('white')

    # ── Volatilidade ─────────────────────────────────────
    ax1.set_facecolor('#fafafa')
    vc = [v if v is not None else 0 for v in vol_c]
    vb = [v if v is not None else 0 for v in vol_b]
    ax1.bar(x - bw/2, vc, bw, label=cfg['name'][:18],
            color=port_col,  alpha=0.85, zorder=3)
    ax1.bar(x + bw/2, vb, bw, label=cfg['bench_short'],
            color=bench_col, alpha=0.85, zorder=3)
    ax1.set_xticks(x); ax1.set_xticklabels(p_labels, fontsize=7)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.1f}%'))
    ax1.tick_params(axis='y', labelsize=7)
    ax1.set_title('Volatilidade Anualizada', fontsize=8.5, fontweight='bold',
                  color='#2d3748', pad=6)
    ax1.set_ylabel('%', fontsize=7, color='#718096')
    ax1.grid(axis='y', color='#e2e8f0', lw=0.5, zorder=0)
    ax1.set_axisbelow(True)
    for sp in ['top','right']: ax1.spines[sp].set_visible(False)
    for sp in ['left','bottom']: ax1.spines[sp].set_color('#e2e8f0')
    ax1.legend(fontsize=6.5, framealpha=0.85, edgecolor='#e2e8f0', fancybox=False,
               loc='upper left')
    # labels on bars
    for xi, (vc_i, vb_i) in enumerate(zip(vc, vb)):
        if vc_i: ax1.text(xi - bw/2, vc_i + 0.2, f'{vc_i:.1f}%', ha='center',
                          fontsize=6, color='#2d3748', va='bottom')
        if vb_i: ax1.text(xi + bw/2, vb_i + 0.2, f'{vb_i:.1f}%', ha='center',
                          fontsize=6, color='#4a5568', va='bottom')

    # ── Máximo Drawdown ───────────────────────────────────
    ax2.set_facecolor('#fafafa')
    dc = [d if d is not None else 0 for d in dd_c]
    db = [d if d is not None else 0 for d in dd_b]
    ax2.bar(x - bw/2, dc, bw, label=cfg['name'][:18],
            color=port_col,  alpha=0.85, zorder=3)
    ax2.bar(x + bw/2, db, bw, label=cfg['bench_short'],
            color=bench_col, alpha=0.85, zorder=3)
    ax2.set_xticks(x); ax2.set_xticklabels(p_labels, fontsize=7)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.1f}%'))
    ax2.tick_params(axis='y', labelsize=7)
    ax2.set_title('Máximo Drawdown', fontsize=8.5, fontweight='bold',
                  color='#2d3748', pad=6)
    ax2.set_ylabel('%', fontsize=7, color='#718096')
    ax2.grid(axis='y', color='#e2e8f0', lw=0.5, zorder=0)
    ax2.set_axisbelow(True)
    for sp in ['top','right']: ax2.spines[sp].set_visible(False)
    for sp in ['left','bottom']: ax2.spines[sp].set_color('#e2e8f0')
    ax2.legend(fontsize=6.5, framealpha=0.85, edgecolor='#e2e8f0', fancybox=False,
               loc='lower left')
    # labels on bars (negative values)
    for xi, (dc_i, db_i) in enumerate(zip(dc, db)):
        if dc_i: ax2.text(xi - bw/2, dc_i - 0.3, f'{dc_i:.1f}%', ha='center',
                          fontsize=6, color='#2d3748', va='top')
        if db_i: ax2.text(xi + bw/2, db_i - 0.3, f'{db_i:.1f}%', ha='center',
                          fontsize=6, color='#4a5568', va='top')

    plt.tight_layout(pad=0.5)
    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _chart_rolling_vol_png(cart_m: dict, bench_m: dict, cfg: dict,
                           window: int = 3,
                           figsize: tuple = (5.6, 2.5),
                           dpi: int = 150) -> bytes:
    """Gráfico de linha: volatilidade anualizada rolling (janela = window meses)."""
    import io as _io
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import math, statistics

    all_m = sorted(set(cart_m) | set(bench_m))
    if len(all_m) < window:
        return b''

    def _rolling(monthly):
        vals = [monthly.get(m) for m in all_m]
        result = []
        for i in range(len(vals)):
            if i < window - 1:
                result.append(None)
                continue
            window_rets = [v / 100 for v in vals[max(0, i - window + 1):i + 1]
                           if v is not None]
            if len(window_rets) < 2:
                result.append(None)
                continue
            std = statistics.stdev(window_rets)
            result.append(round(std * math.sqrt(12) * 100, 2))
        return result

    rv_c = _rolling(cart_m)
    rv_b = _rolling(bench_m)
    labels = [_ml(m) for m in all_m]

    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    ax.set_facecolor('#fafafa'); fig.patch.set_facecolor('white')

    xv = list(range(len(all_m)))
    ax.plot(xv, rv_c, color=cfg['color'],  lw=2,   label=cfg['name'][:18],
            solid_capstyle='round', zorder=3)
    ax.plot(xv, rv_b, color='#a0aec0', lw=1.5, ls='--',
            label=cfg['bench_short'], zorder=2)
    ax.fill_between(xv, [v if v else 0 for v in rv_c], alpha=0.08,
                    color=cfg['color'])

    step = max(1, len(all_m) // 12)
    ax.set_xticks(range(0, len(all_m), step))
    ax.set_xticklabels([labels[i] for i in range(0, len(all_m), step)],
                       fontsize=7.5, rotation=35, ha='right')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.1f}%'))
    ax.tick_params(axis='y', labelsize=7.5)
    ax.set_title(f'Volatilidade Anualizada — Rolling {window}M (%)',
                 fontsize=8.5, fontweight='bold', color='#2d3748', pad=5)
    ax.legend(fontsize=7.5, loc='upper right', framealpha=0.85,
              edgecolor='#e2e8f0', fancybox=False)
    ax.grid(axis='y', color='#e2e8f0', lw=0.5, alpha=0.8, zorder=0)
    ax.set_axisbelow(True)
    for sp in ['top', 'right']: ax.spines[sp].set_visible(False)
    for sp in ['left', 'bottom']: ax.spines[sp].set_color('#e2e8f0')
    ax.set_xlim(-0.5, len(all_m) - 0.5)

    plt.tight_layout(pad=0.4)
    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _chart_drawdown_curve_png(cart_m: dict, bench_m: dict, cfg: dict,
                              figsize: tuple = (5.6, 2.5),
                              dpi: int = 150) -> bytes:
    """Gráfico de linha: curva de drawdown acumulada (underwater chart)."""
    import io as _io
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    all_m = sorted(set(cart_m) | set(bench_m))
    if not all_m:
        return b''

    def _dd_curve(monthly):
        cum, peak = 1.0, 1.0
        result = []
        for m in all_m:
            r = monthly.get(m)
            if r is not None:
                cum *= (1 + r / 100)
            if cum > peak:
                peak = cum
            result.append(round((cum / peak - 1) * 100, 2))
        return result

    dd_c = _dd_curve(cart_m)
    dd_b = _dd_curve(bench_m)
    labels = [_ml(m) for m in all_m]

    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    ax.set_facecolor('#fafafa'); fig.patch.set_facecolor('white')

    xv = list(range(len(all_m)))
    ax.plot(xv, dd_c, color=cfg['color'],  lw=2,   label=cfg['name'][:18],
            solid_capstyle='round', zorder=3)
    ax.plot(xv, dd_b, color='#a0aec0', lw=1.5, ls='--',
            label=cfg['bench_short'], zorder=2)
    ax.fill_between(xv, dd_c, 0, alpha=0.10, color=cfg['color'])
    ax.axhline(0, color='#cbd5e0', lw=0.7, zorder=0)

    step = max(1, len(all_m) // 12)
    ax.set_xticks(range(0, len(all_m), step))
    ax.set_xticklabels([labels[i] for i in range(0, len(all_m), step)],
                       fontsize=7.5, rotation=35, ha='right')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.1f}%'))
    ax.tick_params(axis='y', labelsize=7.5)
    ax.set_title('Máximo Drawdown — Curva Underwater (%)',
                 fontsize=8.5, fontweight='bold', color='#2d3748', pad=5)
    ax.legend(fontsize=7.5, loc='lower left', framealpha=0.85,
              edgecolor='#e2e8f0', fancybox=False)
    ax.grid(axis='y', color='#e2e8f0', lw=0.5, alpha=0.8, zorder=0)
    ax.set_axisbelow(True)
    for sp in ['top', 'right']: ax.spines[sp].set_visible(False)
    for sp in ['left', 'bottom']: ax.spines[sp].set_color('#e2e8f0')
    ax.set_xlim(-0.5, len(all_m) - 0.5)

    plt.tight_layout(pad=0.4)
    buf = _io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _draw_lamina_p3(c, cfg, cart_m, bench_m, last_month=''):
    """Página 3: Volatilidade rolling + Drawdown com tabelas de períodos."""
    import io as _io
    import math, statistics
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import Table, TableStyle

    W, H = A4
    M    = 14 * mm
    accent  = HexColor(cfg['color'])
    aclight = HexColor(cfg['color_light'])

    # Derive reference year from data, not datetime.now()
    _lm = last_month or max(cart_m.keys(), default='')
    cur_year = _lm[:4] if _lm else str(__import__('datetime').datetime.now().year)

    # ── Mini header ──────────────────────────────────────────
    HH2 = 20 * mm
    c.setFillColor(accent)
    c.rect(0, H - HH2, W, HH2, fill=1, stroke=0)
    c.setFillColor(white)

    # Logo
    _p3_logo_bytes = _load_logo_png(LOGO_PDF_PATH)
    _p3_text_x = M
    if _p3_logo_bytes:
        _p3_ir = __import__('reportlab.lib.utils', fromlist=['ImageReader']).ImageReader(
            _io.BytesIO(_p3_logo_bytes))
        _p3_lw, _p3_lh = _p3_ir.getSize()
        _p3_logo_h = 9 * mm
        _p3_logo_w = _p3_logo_h * (_p3_lw / _p3_lh)
        _p3_logo_y = H - HH2 + (HH2 - _p3_logo_h) / 2
        c.drawImage(_p3_ir, M, _p3_logo_y,
                    width=_p3_logo_w, height=_p3_logo_h,
                    preserveAspectRatio=False, mask='auto')
        _p3_text_x = M + _p3_logo_w + 5 * mm

    c.setFont('Lato-Bold', 13)
    c.drawString(_p3_text_x, H - 13 * mm, f'{cfg["name"]} — Risco')
    lm = _ml(_lm)
    c.setFont('Lato', 8)
    c.drawRightString(W - M, H - 13 * mm, lm)

    # ── Layout: duas colunas ──────────────────────────────────
    col_gap = 8 * mm
    col_w   = (W - 2 * M - col_gap) / 2
    y_top   = H - HH2 - 6 * mm

    # ── Helpers de cálculo ────────────────────────────────────
    def _vol_for(monthly, period):
        ms = sorted(monthly.keys())
        if period == 'all':   rets = [monthly[m] / 100 for m in ms]
        elif period == 'ytd': rets = [monthly[m] / 100 for m in ms if m.startswith(cur_year)]
        elif period == '12m': rets = [monthly[m] / 100 for m in ms[-12:]]
        elif period == '3m':  rets = [monthly[m] / 100 for m in ms[-3:]]
        else: rets = []
        if len(rets) < 2: return None
        std = statistics.stdev(rets)
        return round(std * math.sqrt(12) * 100, 2)

    def _dd_for(monthly, period):
        ms = sorted(monthly.keys())
        if period == 'all':   sel = ms
        elif period == 'ytd': sel = [m for m in ms if m.startswith(cur_year)]
        elif period == '12m': sel = ms[-12:]
        elif period == '3m':  sel = ms[-3:]
        else: return None
        if not sel: return None
        cum, peak, dd = 1.0, 1.0, 0.0
        for m in sel:
            cum *= (1 + monthly.get(m, 0) / 100)
            if cum > peak: peak = cum
            d = (cum / peak - 1) * 100
            if d < dd: dd = d
        return round(dd, 2)

    periods = [('ytd', 'YTD'), ('12m', 'Últ. 12M'), ('all', 'Desde Início')]

    # shared table style builder
    def _tbl_style(accent_c, light_c):
        return [
            ('BACKGROUND', (0, 0), (-1, 0), accent_c),
            ('TEXTCOLOR',  (0, 0), (-1, 0), white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Lato-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 6.8),
            ('ALIGN',      (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN',      (0, 1), (0, -1),  'LEFT'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            ('GRID',       (0, 0), (-1, -1), 0.25, HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [HexColor('#ffffff'), light_c]),
            ('FONTNAME',   (0, 1), (0, -1), 'Lato-Bold'),
        ]

    chart_h = 65 * mm   # height of each line chart (2-column layout)
    # figsize matches 2-column display (col_w × chart_h in PDF points)
    _p3_fig_w = col_w / 28.35   # pt → mm → inches (approx)
    _p3_fig_h = chart_h / 28.35
    _p3_fig   = (_p3_fig_w, _p3_fig_h)

    # ════════════════════════════════════════════════
    #  COLUNA ESQUERDA — VOLATILIDADE
    # ════════════════════════════════════════════════
    x_left = M
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(x_left, y_top, 'Volatilidade Anualizada')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.setLineWidth(0.5)
    c.line(x_left, y_top - 2 * mm, x_left + col_w, y_top - 2 * mm)

    vol_png = _chart_rolling_vol_png(cart_m, bench_m, cfg, window=3,
                                     figsize=_p3_fig, dpi=200)
    if vol_png:
        from reportlab.lib.utils import ImageReader as _IR
        c.drawImage(_IR(_io.BytesIO(vol_png)),
                    x_left, y_top - 4 * mm - chart_h,
                    width=col_w, height=chart_h,
                    preserveAspectRatio=False, mask='auto')

    # Tabela de vol por período
    vol_rows = [['Período', cfg['bench_short'], 'Carteira', 'Δ']]
    for pk, pl in periods:
        vc = _vol_for(cart_m,  pk)
        vb = _vol_for(bench_m, pk)
        delta = None if (vc is None or vb is None) else round(vc - vb, 2)
        def _pf(v): return f'{v:.2f}%'.replace('.', ',') if v is not None else '—'
        def _df(v):
            if v is None: return '—'
            s = '+' if v >= 0 else ''
            return f'{s}{v:.2f}%'.replace('.', ',')
        vol_rows.append([pl, _pf(vb), _pf(vc), _df(delta)])

    vtbl_y = y_top - 4 * mm - chart_h - 4 * mm
    vw = col_w / 4
    vtbl = Table(vol_rows, colWidths=[col_w * 0.28] + [col_w * 0.24] * 3)
    vtbl.setStyle(TableStyle(_tbl_style(accent, aclight)))
    _, vh = vtbl.wrapOn(c, col_w, 80)
    vtbl.drawOn(c, x_left, vtbl_y - vh)

    # ════════════════════════════════════════════════
    #  COLUNA DIREITA — DRAWDOWN
    # ════════════════════════════════════════════════
    x_right = M + col_w + col_gap
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(x_right, y_top, 'Máximo Drawdown')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.line(x_right, y_top - 2 * mm, x_right + col_w, y_top - 2 * mm)

    dd_png = _chart_drawdown_curve_png(cart_m, bench_m, cfg,
                                       figsize=_p3_fig, dpi=200)
    if dd_png:
        from reportlab.lib.utils import ImageReader as _IR
        c.drawImage(_IR(_io.BytesIO(dd_png)),
                    x_right, y_top - 4 * mm - chart_h,
                    width=col_w, height=chart_h,
                    preserveAspectRatio=False, mask='auto')

    # Tabela de drawdown por período
    dd_rows = [['Período', cfg['bench_short'], 'Carteira', 'Δ']]
    for pk, pl in periods:
        dc = _dd_for(cart_m,  pk)
        db = _dd_for(bench_m, pk)
        delta = None if (dc is None or db is None) else round(dc - db, 2)
        def _pf(v): return f'{v:.2f}%'.replace('.', ',') if v is not None else '—'
        def _df(v):
            if v is None: return '—'
            s = '+' if v >= 0 else ''
            return f'{s}{v:.2f}%'.replace('.', ',')
        dd_rows.append([pl, _pf(db), _pf(dc), _df(delta)])

    dtbl_y = y_top - 4 * mm - chart_h - 4 * mm
    dtbl = Table(dd_rows, colWidths=[col_w * 0.28] + [col_w * 0.24] * 3)
    dtbl.setStyle(TableStyle(_tbl_style(accent, aclight)))
    _, dh = dtbl.wrapOn(c, col_w, 80)
    dtbl.drawOn(c, x_right, dtbl_y - dh)

    # ── Nota de rodapé ────────────────────────────────────────
    c.setFont('Lato', 5.8)
    c.setFillColor(HexColor('#a0aec0'))
    nota = (f'Volatilidade: desvio padrão dos retornos mensais × √12 (rolling 3M para o gráfico). '
            f'Drawdown: queda do pico ao vale em cada período. '
            f'Eleva MFO — {datetime.now().strftime("%d/%m/%Y")}')
    c.drawCentredString(W / 2, 6 * mm, nota)


def _draw_lamina_p1(c, cfg, cart_m, bench_m, start_date, last_month, dy):
    """Página 1: header + métricas + gráfico + tabela mensal."""
    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors as rlc

    W, H = A4
    M = 14 * mm
    accent  = HexColor(cfg['color'])
    aclight = HexColor(cfg['color_light'])

    cum_cart  = _cumulative(cart_m)
    cum_bench = _cumulative(bench_m)
    last_c  = cart_m.get(last_month, 0)
    last_b  = bench_m.get(last_month, 0)
    tot_c   = list(cum_cart.values())[-1]  if cum_cart  else 0
    tot_b   = list(cum_bench.values())[-1] if cum_bench else 0

    # ── Header band ──────────────────────────────────────────
    HH = 30 * mm
    c.setFillColor(accent)
    c.rect(0, H - HH, W, HH, fill=1, stroke=0)
    c.setFillColor(white)

    # Logo (esquerda do header, centralizado verticalmente)
    _p1_logo_bytes = _load_logo_png(LOGO_PDF_PATH)
    if _p1_logo_bytes:
        import io as _io_p1
        from reportlab.lib.utils import ImageReader as _IRP1
        _p1_ir   = _IRP1(_io_p1.BytesIO(_p1_logo_bytes))
        _p1_lw, _p1_lh = _p1_ir.getSize()
        _p1_logo_h = 13 * mm                            # menor que antes
        _p1_logo_w = _p1_logo_h * (_p1_lw / _p1_lh)
        _p1_logo_y = H - HH + (HH - _p1_logo_h) / 2
        c.drawImage(_p1_ir, M, _p1_logo_y,
                    width=_p1_logo_w, height=_p1_logo_h,
                    preserveAspectRatio=False, mask='auto')

    # Nome da carteira centralizado na página
    c.setFont('Lato-Bold', 17)
    c.drawCentredString(W / 2, H - 13 * mm, cfg['name'])
    c.setFont('Lato', 8.5)
    c.drawCentredString(W / 2, H - 22 * mm, cfg['style'])
    c.setFont('Lato-Bold', 14)
    c.drawRightString(W - M, H - 13 * mm, _ml(last_month))
    c.setFont('Lato', 8)
    c.drawRightString(W - M, H - 22 * mm, 'Factsheet Mensal')

    # ── Metric pills ─────────────────────────────────────────
    y0p  = H - HH - 4 * mm
    ph   = 20 * mm
    gap  = 4 * mm
    pw   = (W - 2 * M - 3 * gap) / 4

    delta_acum = tot_c - tot_b
    pills = [
        ('Retorno no Mês',        _fmt_pct(last_c),        _ml(last_month),        last_c >= 0),
        (f'Δ vs {cfg["bench_short"]} no mês', _fmt_pct(last_c - last_b),
         'no mês',                                                                   (last_c - last_b) >= 0),
        ('Retorno Acumulado',     _fmt_pct(tot_c, 1),      f'desde {start_date}',  tot_c >= 0),
        (f'Δ vs {cfg["bench_short"]} acum.', _fmt_pct(delta_acum, 1),
         f'desde {start_date}',                                                      delta_acum >= 0),
    ]
    for i, (lbl, val, sub, pos) in enumerate(pills):
        x0 = M + i * (pw + gap)
        y1 = y0p - ph
        c.setFillColor(aclight)
        c.roundRect(x0, y1, pw, ph, 4, fill=1, stroke=0)
        c.setFillColor(accent)
        c.rect(x0, y1, 3, ph, fill=1, stroke=0)
        c.setFillColor(HexColor('#718096'))
        c.setFont('Lato', 7)
        c.drawString(x0 + 7, y1 + ph - 6.5, lbl.upper())
        c.setFillColor(HexColor('#276749') if pos else HexColor('#c53030'))
        c.setFont('Lato-Bold', 13)
        c.drawString(x0 + 7, y1 + ph / 2 - 2, val)
        c.setFillColor(HexColor('#a0aec0'))
        c.setFont('Lato', 6.5)
        c.drawString(x0 + 7, y1 + 4.5, sub)

    # ── Performance chart ────────────────────────────────────
    y_chart_top = y0p - ph - 5 * mm
    chart_h = 70 * mm
    png = _chart_perf_png(cart_m, bench_m, cfg)
    if png:
        from reportlab.lib.utils import ImageReader as _IR
        c.drawImage(_IR(_io.BytesIO(png)), M, y_chart_top - chart_h,
                    width=W - 2 * M, height=chart_h,
                    preserveAspectRatio=False, mask='auto')
    c.setFont('Lato', 6.5)
    c.setFillColor(HexColor('#a0aec0'))
    c.drawCentredString(W / 2, y_chart_top - chart_h - 3.5 * mm,
                        f'Retorno acumulado (%) · {cfg["name"]} vs {cfg["bench_label"]} '
                        f'· Base = primeiro fechamento disponível')

    # ── Monthly table ────────────────────────────────────────
    y_tbl = y_chart_top - chart_h - 15 * mm   # gap aumentado
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(M, y_tbl, 'Rentabilidade Mensal')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.setLineWidth(0.5)
    c.line(M, y_tbl - 2 * mm, W - M, y_tbl - 2 * mm)

    years = sorted({m[:4] for m in list(cart_m) + list(bench_m)})
    if not years:
        return

    def _yr(monthly, year):
        r, has = 1.0, False
        for mo in range(1, 13):
            k = f"{year}-{mo:02d}"
            if k in monthly:
                r *= 1 + monthly[k] / 100
                has = True
        return (r - 1) * 100 if has else None

    def _fdiff(a, b):
        if a is None or b is None:
            return '—'
        d = a - b
        s = '+' if d >= 0 else ''
        return f"{s}{d:.2f}%".replace('.', ',')

    # Cumulative return through end of each year, computed from each series directly
    cum_cart_all  = _cumulative(cart_m)
    cum_bench_all = _cumulative(bench_m)

    def _yr_end_cum(cum_dict, y):
        """Last available cumulative value for year y (already in %)."""
        months = [m for m in cum_dict if m.startswith(y)]
        return cum_dict[max(months)] if months else None

    # Display order: most recent year at top, oldest at bottom
    years_display = list(reversed(years))

    COLOR_HDR   = accent
    COLOR_CART  = aclight
    COLOR_BENCH = HexColor('#f7fafc')
    COLOR_DIFF  = HexColor('#fffff0')

    hdr = ['ANO', ''] + _MESES_ABREV + ['No ano', 'Acum.']
    rows = [hdr]
    row_types = []

    for yi, y in enumerate(years_display):
        yc = _yr(cart_m,  y)
        yb = _yr(bench_m, y)
        cc_end = _yr_end_cum(cum_cart_all,  y)
        cb_end = _yr_end_cum(cum_bench_all, y)

        rc = [y, cfg['name'][:20]]
        for mo in range(1, 13):
            rc.append(_fmt_pct(cart_m.get(f"{y}-{mo:02d}")))
        rc += [_fmt_pct(yc), _fmt_pct(cc_end)]
        rows.append(rc)
        row_types.append('cart')

        rb = ['', cfg['bench_short']]
        for mo in range(1, 13):
            rb.append(_fmt_pct(bench_m.get(f"{y}-{mo:02d}")))
        rb += [_fmt_pct(yb), _fmt_pct(cb_end)]
        rows.append(rb)
        row_types.append('bench')

        rd = ['', f'Δ {cfg["bench_short"]}']
        for mo in range(1, 13):
            rd.append(_fdiff(cart_m.get(f"{y}-{mo:02d}"),
                             bench_m.get(f"{y}-{mo:02d}")))
        rd += [_fdiff(yc, yb), _fdiff(cc_end, cb_end)]
        rows.append(rd)
        row_types.append('diff')

    lw = 18 * mm
    lw2 = 24 * mm
    dw = (W - 2 * M - lw - lw2) / 14

    style = [
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_HDR),
        ('TEXTCOLOR',  (0, 0), (-1, 0), white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Lato-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 6),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN',      (0, 0), (1, 0),  'LEFT'),
        ('FONTSIZE',   (0, 1), (-1, -1), 5.8),
        ('ALIGN',      (2, 1), (-1, -1), 'CENTER'),
        ('ALIGN',      (0, 1), (1, -1),  'LEFT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 1.8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.8),
        ('LEFTPADDING',   (0, 0), (-1, -1), 2),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.2, HexColor('#e2e8f0')),
        ('FONTNAME', (-2, 1), (-1, -1), 'Lato-Bold'),
    ]
    for ri, rtype in enumerate(row_types, start=1):
        if rtype == 'cart':
            style += [
                ('BACKGROUND', (0, ri), (-1, ri), COLOR_CART),
                ('FONTNAME',   (1, ri), (1,  ri), 'Lato-Bold'),
            ]
        elif rtype == 'bench':
            style += [
                ('BACKGROUND', (0, ri), (-1, ri), COLOR_BENCH),
                ('TEXTCOLOR',  (1, ri), (1,  ri), HexColor('#718096')),
            ]
        elif rtype == 'diff':
            style += [
                ('BACKGROUND', (0, ri), (-1, ri), COLOR_DIFF),
                ('TEXTCOLOR',  (1, ri), (1,  ri), HexColor('#744210')),
                ('LINEBELOW',  (0, ri), (-1, ri), 0.8, HexColor('#cbd5e0')),
            ]

    # SPAN uses years_display order (same as row building)
    for yi, y in enumerate(years_display):
        rs = 1 + yi * 3
        style += [
            ('SPAN',       (0, rs), (0, rs + 2)),
            ('VALIGN',     (0, rs), (0, rs + 2), 'MIDDLE'),
            ('FONTNAME',   (0, rs), (0, rs),      'Lato-Bold'),
            ('FONTSIZE',   (0, rs), (0, rs),      7.5),
            ('TEXTCOLOR',  (0, rs), (0, rs),      accent),
            ('ALIGN',      (0, rs), (0, rs + 2),  'CENTER'),
            ('BACKGROUND', (0, rs), (0, rs + 2),  COLOR_CART),
        ]

    tbl = Table(rows, colWidths=[lw, lw2] + [dw] * 14, repeatRows=1)
    tbl.setStyle(TableStyle(style))
    avail = y_tbl - 5 * mm - 8 * mm
    _, th = tbl.wrapOn(c, W - 2 * M, avail)
    tbl.drawOn(c, M, max(y_tbl - 5 * mm - th, 8 * mm))


def _draw_lamina_p2(c, cfg, dy, cart_m, bench_m, last_month=''):
    """Página 2: header + pizza de setor + tabela de ativos + seção de risco."""
    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import Table, TableStyle

    W, H = A4
    M = 14 * mm
    accent = HexColor(cfg['color'])

    # Mini header
    HH2 = 20 * mm
    c.setFillColor(accent)
    c.rect(0, H - HH2, W, HH2, fill=1, stroke=0)
    c.setFillColor(white)

    # Logo (esquerda do mini-header)
    _p2_logo_bytes = _load_logo_png(LOGO_PDF_PATH)
    _p2_text_x = M
    if _p2_logo_bytes:
        import io as _io_p2
        from reportlab.lib.utils import ImageReader as _IRP2
        _p2_ir   = _IRP2(_io_p2.BytesIO(_p2_logo_bytes))
        _p2_lw, _p2_lh = _p2_ir.getSize()
        _p2_logo_h = 9 * mm
        _p2_logo_w = _p2_logo_h * (_p2_lw / _p2_lh)
        _p2_logo_y = H - HH2 + (HH2 - _p2_logo_h) / 2
        c.drawImage(_p2_ir, M, _p2_logo_y,
                    width=_p2_logo_w, height=_p2_logo_h,
                    preserveAspectRatio=False, mask='auto')
        _p2_text_x = M + _p2_logo_w + 5 * mm

    lm = _ml(max(cart_m.keys(), default=''))
    c.setFont('Lato-Bold', 13)
    c.drawString(_p2_text_x, H - 13 * mm, f'{cfg["name"]} — Composição e Risco')
    c.setFont('Lato', 8)
    c.drawRightString(W - M, H - 13 * mm, lm)

    y_start = H - HH2 - 6 * mm
    # Left column: pie chart + legend (45% of usable width)
    # Right column: asset table (remaining width with gap)
    left_w  = (W - 2 * M) * 0.45
    gap     = 8 * mm
    right_x = M + left_w + gap
    right_w = W - M - right_x

    # ── Setor (esquerda) ─────────────────────────────────────
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(M, y_start, 'Composição por Setor')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.setLineWidth(0.5)
    c.line(M, y_start - 2 * mm, M + left_w, y_start - 2 * mm)

    # Pie centered in left column
    pie_size = 60 * mm
    pie_x    = M + (left_w - pie_size) / 2
    pie_y    = y_start - pie_size - 5 * mm
    pie_png  = _chart_sector_png(cfg)
    if pie_png:
        from reportlab.lib.utils import ImageReader as _IR
        c.drawImage(_IR(_io.BytesIO(pie_png)), pie_x, pie_y,
                    width=pie_size, height=pie_size,
                    preserveAspectRatio=True, mask='auto')

    # Legend centered below the pie
    sm = {}
    for t, w, s in zip(cfg['tickers'], cfg['weights'], cfg['sectors']):
        sm[s] = sm.get(s, 0) + w
    leg_y = pie_y - 4 * mm
    for sec, pct in sorted(sm.items(), key=lambda x: -x[1]):
        dot_x = M + (left_w - 80) / 2  # center the legend block
        c.setFillColor(HexColor(_SECTOR_COLORS_HEX.get(sec, '#718096')))
        c.roundRect(dot_x, leg_y - 2 * mm, 7, 7, 1.5, fill=1, stroke=0)
        c.setFillColor(HexColor('#2d3748'))
        c.setFont('Lato', 7)
        c.drawString(dot_x + 10, leg_y - 1 * mm, f"{sec}  {pct:.0f}%")
        leg_y -= 11
    legend_bottom = leg_y  # bottom of legend section

    # ── Ativos (direita) ─────────────────────────────────────
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(right_x, y_start, 'Ativos da Carteira')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.line(right_x, y_start - 2 * mm, W - M, y_start - 2 * mm)

    # Only Ticker, Setor, Peso — no DY, no benchmark
    cw = [right_w * 0.22, right_w * 0.53, right_w * 0.25]
    asset_rows = [['Ticker', 'Setor', 'Peso']]
    for t, w, s in zip(cfg['tickers'], cfg['weights'], cfg['sectors']):
        asset_rows.append([t, s, f"{w:.0f}%"])

    astyle = [
        ('BACKGROUND', (0, 0), (-1, 0), accent),
        ('TEXTCOLOR',  (0, 0), (-1, 0), white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Lato-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 7.5),
        ('ALIGN',      (0, 0), (-1, 0),  'CENTER'),
        ('ALIGN',      (2, 1), (-1, -1), 'CENTER'),
        ('ALIGN',      (0, 1), (1, -1),  'LEFT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [HexColor('#ffffff'), HexColor('#f7fafc')]),
        ('FONTNAME',  (0, 1), (0, -1), 'Lato-Bold'),
        ('TEXTCOLOR', (0, 1), (0, -1), accent),
    ]
    avail_h = y_start - 5 * mm - 12 * mm
    atbl = Table(asset_rows, colWidths=cw, repeatRows=1)
    atbl.setStyle(TableStyle(astyle))
    _, ah = atbl.wrapOn(c, right_w, avail_h)
    tbl_top = y_start - 5 * mm
    atbl.drawOn(c, right_x, tbl_top - ah)
    table_bottom = tbl_top - ah  # bottom of asset table

    # ── Seção de Risco (2 colunas) ───────────────────────────
    import math, statistics

    _lm = last_month or max(cart_m.keys(), default='')
    cur_year = _lm[:4] if _lm else str(datetime.now().year)
    aclight = HexColor(cfg['color_light'])

    # Breathing room below composition section
    y_risk = min(table_bottom, legend_bottom) - 14 * mm

    col_gap = 8 * mm
    col_w   = (W - 2 * M - col_gap) / 2

    chart_h   = 58 * mm
    _fig_w    = col_w / 28.35
    _fig_h    = chart_h / 28.35
    _risk_fig = (_fig_w, _fig_h)

    # ── Helpers de cálculo ────────────────────────────────────
    def _vol_for(monthly, period):
        ms = sorted(monthly.keys())
        if period == 'all':   rets = [monthly[m] / 100 for m in ms]
        elif period == 'ytd': rets = [monthly[m] / 100 for m in ms if m.startswith(cur_year)]
        elif period == '12m': rets = [monthly[m] / 100 for m in ms[-12:]]
        elif period == '3m':  rets = [monthly[m] / 100 for m in ms[-3:]]
        else: rets = []
        if len(rets) < 2: return None
        return round(statistics.stdev(rets) * math.sqrt(12) * 100, 2)

    def _dd_for(monthly, period):
        ms = sorted(monthly.keys())
        if period == 'all':   sel = ms
        elif period == 'ytd': sel = [m for m in ms if m.startswith(cur_year)]
        elif period == '12m': sel = ms[-12:]
        elif period == '3m':  sel = ms[-3:]
        else: return None
        if not sel: return None
        cum, peak, dd = 1.0, 1.0, 0.0
        for m in sel:
            cum *= (1 + monthly.get(m, 0) / 100)
            if cum > peak: peak = cum
            d = (cum / peak - 1) * 100
            if d < dd: dd = d
        return round(dd, 2)

    periods = [('ytd', 'YTD'), ('12m', 'Últ. 12M'), ('all', 'Desde Início')]

    def _tbl_style_r(accent_c, light_c):
        return [
            ('BACKGROUND', (0, 0), (-1, 0), accent_c),
            ('TEXTCOLOR',  (0, 0), (-1, 0), white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Lato-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 6.8),
            ('ALIGN',      (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN',      (0, 1), (0, -1),  'LEFT'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            ('GRID',       (0, 0), (-1, -1), 0.25, HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#ffffff'), light_c]),
            ('FONTNAME',   (0, 1), (0, -1), 'Lato-Bold'),
        ]

    # ── Coluna Esquerda — Volatilidade ────────────────────────
    x_left = M
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(x_left, y_risk, 'Volatilidade Anualizada')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.setLineWidth(0.5)
    c.line(x_left, y_risk - 2 * mm, x_left + col_w, y_risk - 2 * mm)

    vol_png = _chart_rolling_vol_png(cart_m, bench_m, cfg, window=3,
                                     figsize=_risk_fig, dpi=200)
    if vol_png:
        from reportlab.lib.utils import ImageReader as _IR2
        c.drawImage(_IR2(_io.BytesIO(vol_png)),
                    x_left, y_risk - 4 * mm - chart_h,
                    width=col_w, height=chart_h,
                    preserveAspectRatio=False, mask='auto')

    vol_rows = [['Período', cfg['bench_short'], 'Carteira', 'Δ']]
    for pk, pl in periods:
        vc = _vol_for(cart_m, pk)
        vb = _vol_for(bench_m, pk)
        delta = None if (vc is None or vb is None) else round(vc - vb, 2)
        def _pf(v): return f'{v:.2f}%'.replace('.', ',') if v is not None else '—'
        def _df(v):
            if v is None: return '—'
            return ('+' if v >= 0 else '') + f'{v:.2f}%'.replace('.', ',')
        vol_rows.append([pl, _pf(vb), _pf(vc), _df(delta)])

    vtbl_y = y_risk - 4 * mm - chart_h - 4 * mm
    vtbl = Table(vol_rows, colWidths=[col_w * 0.28] + [col_w * 0.24] * 3)
    vtbl.setStyle(TableStyle(_tbl_style_r(accent, aclight)))
    _, vh = vtbl.wrapOn(c, col_w, 80)
    vtbl.drawOn(c, x_left, vtbl_y - vh)

    # ── Coluna Direita — Drawdown ─────────────────────────────
    x_right = M + col_w + col_gap
    c.setFont('Lato-Bold', 8.5)
    c.setFillColor(HexColor('#2d3748'))
    c.drawString(x_right, y_risk, 'Máximo Drawdown')
    c.setStrokeColor(HexColor('#e2e8f0'))
    c.line(x_right, y_risk - 2 * mm, x_right + col_w, y_risk - 2 * mm)

    dd_png = _chart_drawdown_curve_png(cart_m, bench_m, cfg,
                                       figsize=_risk_fig, dpi=200)
    if dd_png:
        from reportlab.lib.utils import ImageReader as _IR3
        c.drawImage(_IR3(_io.BytesIO(dd_png)),
                    x_right, y_risk - 4 * mm - chart_h,
                    width=col_w, height=chart_h,
                    preserveAspectRatio=False, mask='auto')

    dd_rows = [['Período', cfg['bench_short'], 'Carteira', 'Δ']]
    for pk, pl in periods:
        dc = _dd_for(cart_m, pk)
        db = _dd_for(bench_m, pk)
        delta = None if (dc is None or db is None) else round(dc - db, 2)
        def _pf(v): return f'{v:.2f}%'.replace('.', ',') if v is not None else '—'
        def _df(v):
            if v is None: return '—'
            return ('+' if v >= 0 else '') + f'{v:.2f}%'.replace('.', ',')
        dd_rows.append([pl, _pf(db), _pf(dc), _df(delta)])

    dtbl_y = y_risk - 4 * mm - chart_h - 4 * mm
    dtbl = Table(dd_rows, colWidths=[col_w * 0.28] + [col_w * 0.24] * 3)
    dtbl.setStyle(TableStyle(_tbl_style_r(accent, aclight)))
    _, dh = dtbl.wrapOn(c, col_w, 80)
    dtbl.drawOn(c, x_right, dtbl_y - dh)

    # ── Rodapé / Disclaimer ───────────────────────────────────
    c.setFont('Lato', 5.8)
    c.setFillColor(HexColor('#a0aec0'))
    disc = (
        'Volatilidade: desvio padrão dos retornos mensais × √12 (rolling 3M para o gráfico). '
        'Drawdown: queda do pico ao vale em cada período. '
        'Material de uso interno — rentabilidade passada não é garantia de resultados futuros. '
        f'Eleva MFO — {datetime.now().strftime("%d/%m/%Y")}'
    )
    c.drawCentredString(W / 2, 6 * mm, disc)


def _ensure_matplotlib():
    """Instala matplotlib automaticamente se não estiver disponível."""
    try:
        import matplotlib  # noqa: F401
    except ImportError:
        print("  ℹ  Instalando 'matplotlib'…")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "matplotlib", "-q"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        print("  ✓  matplotlib instalado.")


def generate_pdf_bytes(portfolio_key: str, data: dict) -> bytes:
    """
    Gera factsheet PDF de 2 páginas para a carteira indicada.
    Retorna os bytes do PDF.
    Chamado pelo serve.py no endpoint /lamina/<portfolio_key>.
    """
    # Garante dependências opcionais instaladas
    _ensure_reportlab()
    _ensure_matplotlib()
    _register_fonts()

    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as _canvas

    cfg = _PORTFOLIO_CFG.get(portfolio_key)
    if cfg is None:
        raise ValueError(f"Portfolio '{portfolio_key}' não encontrado.")

    prices     = data.get('prices', {})
    wh         = data.get('weights_history', {}).get(portfolio_key, {})
    dy         = data.get('dy', {})
    start_date = data.get('start_date', '?')

    cart_m  = _compute_monthly_cart(prices, cfg, wh)
    bench_m = _compute_monthly_bench(prices, cfg['bench'])

    if not cart_m:
        raise ValueError(f"Sem dados de retorno para '{portfolio_key}'. "
                         "Execute o serve.py para atualizar o dashboard_data.json.")

    # Cap to last closed month — never show partial current-month data
    _tnow  = datetime.now()
    _tprev = f"{_tnow.year}-{_tnow.month-1:02d}" if _tnow.month > 1 else f"{_tnow.year-1}-12"
    last_month = min(max(cart_m.keys()), _tprev)

    # Filter series so no chart/table contains months beyond last_month
    cart_m  = {k: v for k, v in cart_m.items()  if k <= last_month}
    bench_m = {k: v for k, v in bench_m.items() if k <= last_month}

    buf = _io.BytesIO()
    c = _canvas.Canvas(buf, pagesize=A4)

    _draw_lamina_p1(c, cfg, cart_m, bench_m, start_date, last_month, dy)
    c.showPage()
    _draw_lamina_p2(c, cfg, dy, cart_m, bench_m, last_month)
    c.showPage()
    c.save()

    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    main()
